#!/usr/bin/env python3
"""
TCN Campaign Report Builder
============================
Ingests a raw TCN export workbook (one tab of Outbound calls, one tab of Inbound
calls), regenerates the EST/IST time + 2-hour bucket columns from the raw
timestamps, runs the full OB/IB analysis (best-time pivots, outcome-by-time
flags, connectivity, cost), and writes a clean formatted report workbook.

Designed to be CLIENT-AGNOSTIC and PERIOD-AGNOSTIC:
  - Works for any client (Prime Recovery today, others later).
  - Works for daily / weekly / monthly files (period is parsed from the filename
    or passed explicitly); the analysis logic is identical, only labels change.

Usage:
    python tcn_report.py "<input.xlsx>" [--out "<output.xlsx>"]
                                        [--client "Prime Recovery"]
                                        [--period daily|weekly|monthly]

If --client / --period are omitted they are parsed from the filename when it
follows the convention:  Client_Period_<date(s)>.xlsx
e.g.  PrimeRecovery_Daily_2026-06-24.xlsx
      PrimeRecovery_Weekly_2026-06-16_to_2026-06-22.xlsx
      PrimeRecovery_Monthly_2026-06.xlsx
"""

import argparse
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# --------------------------------------------------------------------------- #
# CONFIG  --  edit here to onboard a new client or tweak business rules        #
# --------------------------------------------------------------------------- #
IST_OFFSET_FROM_EST = timedelta(hours=10, minutes=30)   # EST(-05:00) -> IST(+05:30)

# How we recognise the two raw tabs (case-insensitive substring match, in order)
OB_SHEET_HINTS = ["ob data", "outbound", "ob"]
IB_SHEET_HINTS = ["ib data", "inbound", "ib"]

# Column resolution: logical name -> list of acceptable header variants
OB_COLS = {
    "timestamp":      ["Start Time", "Start Date Time", "Call Date-Time"],
    "result":         ["Result"],
    "last_element":   ["Last Element"],
    "linkback_len":   ["Linkback Length"],
    "total_len":      ["Total Length"],
    "delivery_cost":  ["Delivery Cost"],
    "linkback_cost":  ["Linkback Cost"],
    "total_cost":     ["Total Cost"],
    # agent + disposition columns (added to the OB dump from 2026-06-29 on; the
    # same layer we already run for inbound — analysis is no-op if absent)
    "talk_dur":       ["Agent Call Talk Duration"],
    "hold_dur":       ["Agent Call Hold Duration"],
    "wrap_dur":       ["Agent Call Wrap up Duration"],
    "agent_first":    ["Agent First Name"],
    "agent_last":     ["Agent Last Name"],
    "disposition":    ["Agent Dispositions", "An Agent Call Response",
                       "Agent Call Response", "Agent Disposition"],
}
IB_COLS = {
    "timestamp":      ["Call Date-Time", "Start Time"],
    "result":         ["Result"],
    "last_element":   ["Last Element"],
    "linkback_len":   ["Linkback Length"],
    "total_len":      ["Total Length"],
    "total_cost":     ["Total Cost"],
    "inbound_cost":   ["Inbound Cost"],
    "talk_dur":       ["Agent Call Talk Duration"],
    "hold_dur":       ["Agent Call Hold Duration"],
    "wrap_dur":       ["Agent Call Wrap up Duration"],
    "agent_first":    ["Agent First Name"],
    "agent_last":     ["Agent Last Name"],
    "disposition":    ["An Agent Call Response", "Agent Call Response",
                       "Agent Dispositions", "Agent Disposition"],
}

# Business definitions (kept as data, not hard rules in code, so they're visible)
OB_ANSWERED_LINKCALL = "Answered Linkcall"
IB_CONNECTED_RESULTS = ["Answered Linkcall", "Answered"]   # IB "connectivity"

# Inbound agent-disposition ("An Agent Call Response") taxonomy. The codes are
# structured hyphen-segments: <contact-class>-<party>-<outcome...>, e.g.
#   C-RPC-PTP        -> Connected, Right-Party Contact, Promise To Pay
#   C-RPC-DNC-NU     -> Connected, Right-Party, Do-Not-Call / Not Usable
#   C-TPC-WRONG      -> Connected, Third-Party, Wrong number
#   NC-DISCONNECT    -> Not Connected, disconnected
# We parse the segments so drill-downs (by contact class / party / outcome) come
# for free without hard-coding the full code list. "Connected" = code starts C-.
AGENT_BLANK_LABEL = "System / Unassigned"
DISP_CLASS = {"C": "Connected", "NC": "Not Connected"}
DISP_PARTY = {"RPC": "Right Party (RPC)", "TPC": "Third Party (TPC)"}


def parse_disposition(code):
    """Split an agent-response code into (class, party, outcome) labels.
    Returns a dict; unknown/blank codes degrade gracefully."""
    raw = "" if code is None else str(code).strip()
    if not raw or raw.lower() in ("nan", "na", "none"):
        return {"raw": None, "class": None, "party": None, "outcome": None,
                "connected": None}
    segs = [s for s in raw.split("-") if s != ""]
    cls = DISP_CLASS.get(segs[0], segs[0]) if segs else None
    connected = (segs[0] == "C") if segs else None
    party, outcome = None, None
    if len(segs) >= 2:
        if segs[1] in DISP_PARTY:
            party = DISP_PARTY[segs[1]]
            outcome = "-".join(segs[2:]) or None
        else:
            outcome = "-".join(segs[1:]) or None
    return {"raw": raw, "class": cls, "party": party, "outcome": outcome,
            "connected": connected}


def agent_name_series(df, m):
    """Combined 'First Last' agent name per row; blanks -> AGENT_BLANK_LABEL."""
    fcol, lcol = m.get("agent_first"), m.get("agent_last")
    if not fcol and not lcol:
        return None
    first = df[fcol].astype(str) if fcol else ""
    last = df[lcol].astype(str) if lcol else ""

    def _clean(s):
        s = "" if s is None else str(s).strip()
        return "" if s.lower() in ("nan", "none", "na") else s

    names = []
    for fi, la in zip(
            (first if fcol else [""] * len(df)),
            (last if lcol else [""] * len(df))):
        full = (f"{_clean(fi)} {_clean(la)}").strip()
        names.append(full if full else AGENT_BLANK_LABEL)
    return pd.Series(names, index=df.index)

# --------------------------------------------------------------------------- #
# Styling                                                                      #
# --------------------------------------------------------------------------- #
FONT_NAME = "Arial"
NAVY = "1F3864"; BLUE = "2E5496"; LIGHT = "D9E1F2"; LIGHT2 = "EEF3FB"
GREEN = "548235"; AMBER = "BF8F00"; RED = "C00000"; GREY = "808080"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def f(sz=10, bold=False, color="000000"):
    return Font(name=FONT_NAME, size=sz, bold=bold, color=color)


def fill(hexc):
    return PatternFill("solid", fgColor=hexc)


# --------------------------------------------------------------------------- #
# Time helpers                                                                 #
# --------------------------------------------------------------------------- #
TS_RE = re.compile(r"(\d{4}-\d{2}-\d{2})[ T](\d{2}:\d{2}:\d{2})")

BUCKET_ORDER = [
    "12 AM - 2 AM", "2 AM - 4 AM", "4 AM - 6 AM", "6 AM - 8 AM",
    "8 AM - 10 AM", "10 AM - 12 PM", "12 PM - 2 PM", "2 PM - 4 PM",
    "4 PM - 6 PM", "6 PM - 8 PM", "8 PM - 10 PM", "10 PM - 12 AM",
]


def _h12(h):
    ampm = "AM" if h < 12 else "PM"
    hr = h % 12
    if hr == 0:
        hr = 12
    return f"{hr} {ampm}"


def bucket_label(hour):
    start = hour - (hour % 2)
    end = (start + 2) % 24
    return f"{_h12(start)} - {_h12(end)}"


def parse_est(raw):
    """Return a datetime carrying the wall-clock EST time-of-day (offset ignored,
    matching the client's existing methodology)."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    m = TS_RE.search(str(raw))
    if not m:
        return None
    return datetime.strptime(f"{m.group(1)} {m.group(2)}", "%Y-%m-%d %H:%M:%S")


def derive_time_columns(df, ts_col):
    est = df[ts_col].apply(parse_est)
    df["_est_dt"] = est
    df["EST Date"] = est.apply(lambda d: d.date() if d else None)
    df["EST Time"] = est.apply(lambda d: d.strftime("%H:%M:%S") if d else None)
    df["EST Bucket"] = est.apply(lambda d: bucket_label(d.hour) if d else None)
    ist = est.apply(lambda d: (d + IST_OFFSET_FROM_EST) if d else None)
    df["IST"] = ist.apply(lambda d: d.strftime("%H:%M:%S") if d else None)
    df["IST Bucket"] = ist.apply(lambda d: bucket_label(d.hour) if d else None)
    return df


# --------------------------------------------------------------------------- #
# Load + normalise                                                             #
# --------------------------------------------------------------------------- #
def pick_sheet(sheets, hints):
    names = {n.lower(): n for n in sheets}
    for hint in hints:
        for low, orig in names.items():
            if hint == low:
                return orig
    for hint in hints:
        for low, orig in names.items():
            if hint in low:
                return orig
    return None


def resolve(df, spec, required=()):
    """Map logical names -> actual df columns; return rename dict."""
    out = {}
    lower = {c.lower().strip(): c for c in df.columns}
    for logical, variants in spec.items():
        for v in variants:
            if v.lower().strip() in lower:
                out[logical] = lower[v.lower().strip()]
                break
    missing = [r for r in required if r not in out]
    if missing:
        raise ValueError(f"Missing required columns {missing}. Found: {list(df.columns)}")
    return out


def load(input_path):
    xls = pd.read_excel(input_path, sheet_name=None)
    ob_name = pick_sheet(xls.keys(), OB_SHEET_HINTS)
    ib_name = pick_sheet(xls.keys(), IB_SHEET_HINTS)
    ob = xls[ob_name].copy() if ob_name else None
    ib = xls[ib_name].copy() if ib_name else None

    if ob is not None:
        m = resolve(ob, OB_COLS, required=("timestamp", "result"))
        ob = derive_time_columns(ob, m["timestamp"])
        ob.attrs["meta"] = m
    if ib is not None:
        m = resolve(ib, IB_COLS, required=("timestamp", "result"))
        ib = derive_time_columns(ib, m["timestamp"])
        ib.attrs["meta"] = m
    return ob, ib


# --------------------------------------------------------------------------- #
# Analysis                                                                     #
# --------------------------------------------------------------------------- #
def pct(n, d):
    return (n / d) if d else 0.0


def bucket_series(df, mask=None):
    """Counts per bucket for EST and IST, ordered, only buckets present kept."""
    sub = df[mask] if mask is not None else df
    est = sub["EST Bucket"].value_counts().reindex(BUCKET_ORDER).dropna()
    ist = sub["IST Bucket"].value_counts().reindex(BUCKET_ORDER).dropna()
    return est.astype(int), ist.astype(int)


def outcome_by_bucket(df, result_col, bucket_col):
    """Crosstab: rows=bucket (ordered), cols=result, values=count."""
    ct = pd.crosstab(df[bucket_col], df[result_col])
    ct = ct.reindex([b for b in BUCKET_ORDER if b in ct.index])
    ct["Total"] = ct.sum(axis=1)
    return ct


# --------------------------------------------------------------------------- #
# Day-of-week / week-block breakdowns (used by weekly + monthly reports)       #
# --------------------------------------------------------------------------- #
WEEKDAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday",
                 "Friday", "Saturday", "Sunday"]
# Month split per client request: 1-7, 8-14, 15-21, 22-end.
WEEKBLOCK_ORDER = ["Week 1 (1-7)", "Week 2 (8-14)",
                   "Week 3 (15-21)", "Week 4 (22-end)"]


def week_block(day):
    """Map a day-of-month (1-31) to its week block label."""
    if day <= 7:
        return WEEKBLOCK_ORDER[0]
    if day <= 14:
        return WEEKBLOCK_ORDER[1]
    if day <= 21:
        return WEEKBLOCK_ORDER[2]
    return WEEKBLOCK_ORDER[3]


def _rate_frame(keys, success_mask, order):
    """Group rows by `keys`, returning a frame indexed in `order` with columns
    total / success / rate. Keys that are None are dropped."""
    g = pd.DataFrame({"key": list(keys), "ok": list(success_mask.astype(int))})
    g = g[g["key"].notna()]
    if g.empty:
        return pd.DataFrame(columns=["total", "success", "rate"])
    agg = g.groupby("key")["ok"].agg(total="size", success="sum")
    agg = agg.reindex([k for k in order if k in agg.index])
    agg["total"] = agg["total"].astype(int)
    agg["success"] = agg["success"].astype(int)
    agg["rate"] = [pct(s, t) for s, t in zip(agg["success"], agg["total"])]
    return agg


def period_breakdowns(df, success_mask):
    """Day-of-week, week-block, and best-day-per-week breakdowns for one stream.
    `success` = answered linkcalls (OB) or connected calls (IB)."""
    dt = df["_est_dt"]
    weekday = dt.apply(lambda d: d.strftime("%A") if d is not None else None)
    blk = dt.apply(lambda d: week_block(d.day) if d is not None else None)

    dow = _rate_frame(weekday, success_mask, WEEKDAY_ORDER)
    weekblock = _rate_frame(blk, success_mask, WEEKBLOCK_ORDER)

    # best day within each week block (for monthly "best day of each week")
    block_dow = {}
    tmp = pd.DataFrame({"blk": list(blk), "wd": list(weekday),
                        "ok": list(success_mask.astype(int))})
    tmp = tmp[tmp["blk"].notna() & tmp["wd"].notna()]
    for b in WEEKBLOCK_ORDER:
        sub = tmp[tmp["blk"] == b]
        if sub.empty:
            continue
        agg = sub.groupby("wd")["ok"].agg(total="size", success="sum")
        agg = agg.reindex([d for d in WEEKDAY_ORDER if d in agg.index])
        agg["total"] = agg["total"].astype(int)
        agg["success"] = agg["success"].astype(int)
        agg["rate"] = [pct(s, t) for s, t in zip(agg["success"], agg["total"])]
        block_dow[b] = agg
    return {"dow": dow, "weekblock": weekblock, "block_dow": block_dow}


def best_by_rate(frame, want="best"):
    """Return (label, success, total, rate) for the highest (or lowest) rate row.
    Rows with zero calls are ignored. Returns None if nothing qualifies."""
    if frame is None or len(frame) == 0:
        return None
    f2 = frame[frame["total"] > 0]
    if f2.empty:
        return None
    idx = f2["rate"].idxmax() if want == "best" else f2["rate"].idxmin()
    row = f2.loc[idx]
    return (idx, int(row["success"]), int(row["total"]), float(row["rate"]))


# --------------------------------------------------------------------------- #
# Connect-RATE analytics (the leadership-planning layer)                        #
# --------------------------------------------------------------------------- #
def rate_by_bucket(df, success_mask, bucket_col):
    """Connect rate per 2-hour window: total / success / rate, ordered."""
    return _rate_frame(df[bucket_col], success_mask, BUCKET_ORDER)


def day_hour_matrix(df, success_mask):
    """Weekday x EST-window matrices: dial totals, successes, and connect rate.
    The rate matrix powers the heatmap leaders use to pick call slots."""
    wd = df["_est_dt"].apply(lambda d: d.strftime("%A") if d is not None else None)
    tmp = pd.DataFrame({"wd": list(wd), "bk": list(df["EST Bucket"]),
                        "ok": list(success_mask.astype(int))})
    tmp = tmp.dropna(subset=["wd", "bk"])
    if tmp.empty:
        empty = pd.DataFrame()
        return empty, empty, empty
    totals = tmp.pivot_table(index="wd", columns="bk", values="ok",
                             aggfunc="size", fill_value=0)
    succ = tmp.pivot_table(index="wd", columns="bk", values="ok",
                           aggfunc="sum", fill_value=0)
    rows = [d for d in WEEKDAY_ORDER if d in totals.index]
    cols = [b for b in BUCKET_ORDER if b in totals.columns]
    totals = totals.reindex(index=rows, columns=cols, fill_value=0)
    succ = succ.reindex(index=rows, columns=cols, fill_value=0)
    rate = (succ / totals.replace(0, pd.NA))
    return totals, succ, rate


def peak_windows(rate_frame, min_share=0.04, top=3):
    """Top windows ranked by connect rate, but only those carrying a meaningful
    share of volume (default >=4% of dials) so we don't recommend a 1-dial fluke.
    Returns list of (window, success, total, rate)."""
    if rate_frame is None or len(rate_frame) == 0:
        return []
    grand = rate_frame["total"].sum()
    if not grand:
        return []
    cut = max(1, int(grand * min_share))
    elig = rate_frame[rate_frame["total"] >= cut]
    if elig.empty:
        elig = rate_frame[rate_frame["total"] > 0]
    elig = elig.sort_values("rate", ascending=False)
    return [(idx, int(r["success"]), int(r["total"]), float(r["rate"]))
            for idx, r in elig.head(top).iterrows()]


def cost_per_connect_by_bucket(cost_by_est, conn_by_est):
    """$ spent / connects, per EST window. Aligns two Series on bucket label."""
    if cost_by_est is None or conn_by_est is None:
        return None
    out = {}
    for b in BUCKET_ORDER:
        c = float(cost_by_est.get(b, 0)) if b in getattr(cost_by_est, "index", []) else 0.0
        k = int(conn_by_est.get(b, 0)) if b in getattr(conn_by_est, "index", []) else 0
        if k > 0:
            out[b] = c / k
    return pd.Series(out) if out else None


def analyze_ob(ob):
    m = ob.attrs["meta"]
    total = len(ob)
    rc = ob[m["result"]].value_counts()
    al_mask = ob[m["result"]] == OB_ANSWERED_LINKCALL
    al = int(al_mask.sum())
    est_al, ist_al = bucket_series(ob, al_mask)
    # non-answered-linkcall = "everything else", flagged by bucket
    non_mask = ~al_mask
    res = {
        "total": total,
        "answered_linkcalls": al,
        "connectivity": pct(al, total),
        "result_counts": rc,
        "est_al": est_al, "ist_al": ist_al,
        "outcomes_est": outcome_by_bucket(ob, m["result"], "EST Bucket"),
        "outcomes_ist": outcome_by_bucket(ob, m["result"], "IST Bucket"),
        "non_est": bucket_series(ob, non_mask)[0],
        "non_ist": bucket_series(ob, non_mask)[1],
        "non_total": int(non_mask.sum()),
    }
    res.update(period_breakdowns(ob, al_mask))
    # connect-rate layer
    res["rate_est"] = rate_by_bucket(ob, al_mask, "EST Bucket")
    res["rate_ist"] = rate_by_bucket(ob, al_mask, "IST Bucket")
    res["heat_tot"], res["heat_succ"], res["heat_rate"] = day_hour_matrix(ob, al_mask)
    res["dials_per_connect"] = (total / al) if al else None
    # waste: non-answered outcomes ranked by volume (where dial budget leaks)
    res["waste_counts"] = rc[rc.index != OB_ANSWERED_LINKCALL]
    # cost
    for k in ("delivery_cost", "linkback_cost", "total_cost"):
        col = m.get(k)
        res[k] = float(pd.to_numeric(ob[col], errors="coerce").fillna(0).sum()) if col else None
    if m.get("total_cost"):
        tc = pd.to_numeric(ob[m["total_cost"]], errors="coerce").fillna(0)
        res["cost_per_dial"] = pct(tc.sum(), total)
        res["cost_per_al"] = pct(tc.sum(), al)
        res["cost_by_result"] = tc.groupby(ob[m["result"]]).sum().sort_values(ascending=False)
        res["cost_by_est"] = tc.groupby(ob["EST Bucket"]).sum().reindex(
            [b for b in BUCKET_ORDER if b in set(ob["EST Bucket"])]).dropna()
        # spend that produced a connect vs spend that didn't
        cost_on_al = float(res["cost_by_result"].get(OB_ANSWERED_LINKCALL, 0.0))
        res["cost_connected"] = cost_on_al
        res["cost_wasted"] = float(tc.sum()) - cost_on_al
        res["cost_per_connect_est"] = cost_per_connect_by_bucket(
            res["cost_by_est"], res["est_al"])
    # agent + disposition layer (Col K/L + "An Agent Call Response"), same as IB.
    # Connected (C-*) here is the agent-response classification, independent of
    # the outbound "Answered Linkcall" connectivity above.
    res["agentdisp"] = analyze_ib_agents(ob, m)
    res["_df"] = ob          # kept for the live-formula raw-data sheet
    return res


def analyze_ib_agents(ib, m):
    """Agent-disposition analytics for inbound. Returns None when the dump has
    no disposition column at all. Builds:
      - disposition distribution (raw codes) + class/party/outcome rollups
      - per-agent table (handled, connected, connect-rate, PTP, payment, talk)
    Connected = response code starts with 'C-'. Blank-agent rows (abandoned /
    voicemail / not connected) are kept in totals under AGENT_BLANK_LABEL but
    excluded from the per-agent ranking."""
    if not m.get("disposition"):
        return None
    disp_raw = ib[m["disposition"]]
    parsed = disp_raw.apply(parse_disposition)
    has_disp = parsed.apply(lambda p: p["raw"] is not None)
    if not bool(has_disp.any()):
        return None

    agents = agent_name_series(ib, m)
    talk = (pd.to_numeric(ib[m["talk_dur"]], errors="coerce")
            if m.get("talk_dur") else pd.Series([None] * len(ib), index=ib.index))

    w = pd.DataFrame({
        "agent": agents if agents is not None else AGENT_BLANK_LABEL,
        "raw": parsed.apply(lambda p: p["raw"]),
        "cls": parsed.apply(lambda p: p["class"]),
        "party": parsed.apply(lambda p: p["party"]),
        "outcome": parsed.apply(lambda p: p["outcome"]),
        "connected": parsed.apply(lambda p: p["connected"]),
        "talk": talk,
    })
    d = w[has_disp.values].copy()           # only dispositioned rows
    disp_total = int(len(d))
    disp_counts = d["raw"].value_counts()
    disp_connected = int((d["connected"] == True).sum())  # noqa: E712

    def _counts(col):
        return d[col].dropna().value_counts()

    res = {
        "disp_total": disp_total,
        "disp_counts": disp_counts,
        "disp_connected": disp_connected,
        "disp_connected_rate": pct(disp_connected, disp_total),
        "by_class": _counts("cls"),
        "by_party": _counts("party"),
        "by_outcome": _counts("outcome"),
    }

    # ---- per-agent table (exclude the blank/system bucket from ranking) ----
    agent_rows = []
    for name, g in d.groupby("agent"):
        handled = int(len(g))
        conn = int((g["connected"] == True).sum())  # noqa: E712
        out = g["outcome"]
        ptp = int((out == "PTP").sum())
        payment = int((out == "PAYMENT").sum())
        no_ptp = int((out == "NO_PTP").sum())
        tk = pd.to_numeric(g["talk"], errors="coerce").dropna()
        top = g["raw"].value_counts()
        agent_rows.append({
            "agent": name, "handled": handled, "connected": conn,
            "rate": pct(conn, handled), "ptp": ptp, "payment": payment,
            "no_ptp": no_ptp,
            "avg_talk": float(tk.mean()) if len(tk) else 0.0,
            "top_disp": top.index[0] if len(top) else "—",
        })
    ranked = sorted([r for r in agent_rows if r["agent"] != AGENT_BLANK_LABEL],
                    key=lambda r: (-r["handled"], -r["rate"]))
    # System / Unassigned = all rows with no agent name (abandoned, voicemail,
    # not connected). These carry no disposition, so they come from the FULL
    # frame, not the dispositioned subset.
    blank_handled = int((w["agent"] == AGENT_BLANK_LABEL).sum())
    res["agents"] = ranked
    res["agent_blank"] = ({"agent": AGENT_BLANK_LABEL, "handled": blank_handled}
                          if blank_handled else None)
    res["n_agents"] = len(ranked)
    return res


def analyze_ib(ib):
    m = ib.attrs["meta"]
    total = len(ib)
    rc = ib[m["result"]].value_counts()
    conn_mask = ib[m["result"]].isin(IB_CONNECTED_RESULTS)
    conn = int(conn_mask.sum())
    est_c, ist_c = bucket_series(ib, conn_mask)
    miss_mask = ~conn_mask
    res = {
        "total": total,
        "connected": conn,
        "connectivity": pct(conn, total),
        "result_counts": rc,
        "est_conn": est_c, "ist_conn": ist_c,
        "outcomes_est": outcome_by_bucket(ib, m["result"], "EST Bucket"),
        "outcomes_ist": outcome_by_bucket(ib, m["result"], "IST Bucket"),
        "miss_est": bucket_series(ib, miss_mask)[0],
        "miss_ist": bucket_series(ib, miss_mask)[1],
        "miss_total": int(miss_mask.sum()),
    }
    res.update(period_breakdowns(ib, conn_mask))
    # connect-rate + missed-rate layer (drives the inbound staffing signal)
    res["rate_est"] = rate_by_bucket(ib, conn_mask, "EST Bucket")
    res["rate_ist"] = rate_by_bucket(ib, conn_mask, "IST Bucket")
    res["heat_tot"], res["heat_succ"], res["heat_rate"] = day_hour_matrix(ib, conn_mask)
    res["miss_rate"] = pct(int(miss_mask.sum()), total)
    res["miss_by_est"] = rate_by_bucket(ib, miss_mask, "EST Bucket")  # 'success' col = missed
    for k in ("total_cost", "inbound_cost"):
        col = m.get(k)
        res[k] = float(pd.to_numeric(ib[col], errors="coerce").fillna(0).sum()) if col else None
    # agent durations on connected/linkcall calls
    if m.get("talk_dur"):
        td = pd.to_numeric(ib[m["talk_dur"]], errors="coerce").dropna()
        res["avg_talk"] = float(td.mean()) if len(td) else 0.0
        res["total_talk"] = float(td.sum()) if len(td) else 0.0
    # agent + disposition layer (Col K/L + "An Agent Call Response")
    res["agentdisp"] = analyze_ib_agents(ib, m)
    res["_df"] = ib          # kept for the live-formula raw-data sheet
    return res


# --------------------------------------------------------------------------- #
# Recommendations  --  turn the numbers into plain-English leadership actions  #
# --------------------------------------------------------------------------- #
def _fmt_windows(wins):
    return ", ".join(f"{w} ({r:.0%})" for w, _s, _t, r in wins)


def recommendations(period, ob_res, ib_res):
    """Plain-English, decision-ready bullets for the Leadership View."""
    recs = []
    if ob_res and ob_res["answered_linkcalls"]:
        peaks = peak_windows(ob_res.get("rate_est"))
        if peaks:
            recs.append(
                "Concentrate outbound dialing in " + _fmt_windows(peaks)
                + " EST — these are the highest connect-rate windows with real volume.")
        worst = best_by_rate(ob_res.get("rate_est"), want="worst")
        if peaks and worst and worst[0] != peaks[0][0]:
            best_r = peaks[0][3]
            if worst[3] > 0 and best_r / worst[3] >= 1.3:
                recs.append(
                    f"Pull dials out of {worst[0]} EST ({worst[3]:.0%} connect) — "
                    f"your best window connects {best_r / worst[3]:.1f}x more often.")
        dpc = ob_res.get("dials_per_connect")
        if dpc:
            recs.append(
                f"At the current {ob_res['connectivity']:.1%} connect rate it takes "
                f"~{dpc:.0f} dials per connect — size dial lists to your connect target "
                f"(e.g. {int(round(dpc)) * 100:,} dials for 100 connects).")
        cpc = ob_res.get("cost_per_connect_est")
        if cpc is not None and len(cpc):
            cheap = cpc.idxmin()
            recs.append(
                f"Cheapest connects come from {cheap} EST at ${cpc.min():,.3f}/connect "
                f"vs ${cpc.max():,.3f} in the priciest window — weight budget toward the cheap end.")
        wc = ob_res.get("waste_counts")
        if wc is not None and len(wc) and ob_res["total"]:
            top_waste = wc.index[0]
            recs.append(
                f"Biggest dial leak: '{top_waste}' is {wc.iloc[0] / ob_res['total']:.0%} of all dials "
                f"({int(wc.iloc[0]):,}) — review list quality / dial caps to cut it.")
        if period in ("weekly", "monthly"):
            bd = best_by_rate(ob_res.get("dow"))
            wd = best_by_rate(ob_res.get("dow"), want="worst")
            if bd:
                msg = f"Best outbound day is {bd[0]} ({bd[3]:.1%})"
                if wd and wd[0] != bd[0]:
                    msg += (f"; {wd[0]} is weakest ({wd[3]:.1%}) — concentrate volume "
                            f"on the strong days and trim spend on the weak ones")
                recs.append(msg + ".")
        if period == "monthly":
            bw = best_by_rate(ob_res.get("weekblock"))
            if bw:
                recs.append(f"Strongest week of the month: {bw[0]} ({bw[3]:.1%}) — "
                            "front-load high-value lists there.")
    if ib_res and ib_res["total"]:
        mr = ib_res.get("miss_rate", 0)
        mb = ib_res.get("miss_by_est")
        peak_miss = best_by_rate(mb) if mb is not None else None  # 'success'=missed here
        if mr >= 0.10 and peak_miss:
            recs.append(
                f"Inbound: {mr:.0%} of calls go unanswered, peaking in {peak_miss[0]} EST "
                f"({peak_miss[1]} missed) — add agent coverage in that window to recover leads.")
        elif mr:
            recs.append(f"Inbound miss rate is {mr:.0%} — healthy; hold current staffing.")
        if period in ("weekly", "monthly"):
            wd = best_by_rate(ib_res.get("dow"), want="worst")
            if wd:
                recs.append(f"Inbound connects worst on {wd[0]} ({wd[3]:.1%}) — "
                            "check staffing/coverage that day.")
    if not recs:
        recs.append("Not enough volume in this period to make confident recommendations.")
    return recs


# --------------------------------------------------------------------------- #
# Live-formula raw-data layer                                                  #
# --------------------------------------------------------------------------- #
# Every analytical number in the report is written as a VISIBLE Excel formula
# (COUNTIF / COUNTIFS / SUMIFS / AVERAGEIFS / cell-ratios) that points at a copy
# of the raw dump embedded as a sheet (OB_RawData / IB_RawData). Leadership can
# click any cell and see exactly how the value is derived. We still compute the
# same numbers in pandas (for the trend history + alerts + row ordering/labels);
# only the cell VALUES become formulas so Excel recomputes them live.

# Helper columns we add to the raw sheet so the formulas stay simple + readable.
H_WEEKDAY = "Weekday"
H_WEEKBLOCK = "Week Block"
H_ISCONNECT = "Is Connect"          # Yes/No  (OB: Answered Linkcall; IB: connected results)
H_AGENT = "Agent Name"
H_DISP_CODE = "Disp Code"
H_HASDISP = "Has Disp"              # Yes/No
H_DISP_CLASS = "Disp Class"
H_DISP_PARTY = "Disp Party"
H_DISP_OUTCOME = "Disp Outcome"
H_DISP_CONN = "Disp Connected"     # Yes/No/""


def _prep_raw(df, meta, stream):
    """Return a copy of the raw frame with derived helper columns appended and
    the internal datetime column dropped, ready to dump as the *_RawData sheet."""
    d = df.copy()
    dt = d["_est_dt"]
    d[H_WEEKDAY] = dt.apply(lambda x: x.strftime("%A") if x is not None else None)
    d[H_WEEKBLOCK] = dt.apply(lambda x: week_block(x.day) if x is not None else None)
    res = d[meta["result"]]
    conn = (res == OB_ANSWERED_LINKCALL) if stream == "OB" else res.isin(IB_CONNECTED_RESULTS)
    d[H_ISCONNECT] = ["Yes" if b else "No" for b in conn]
    if meta.get("disposition"):
        parsed = d[meta["disposition"]].apply(parse_disposition)
        an = agent_name_series(d, meta)
        d[H_AGENT] = an if an is not None else AGENT_BLANK_LABEL
        d[H_DISP_CODE] = parsed.apply(lambda p: p["raw"])
        d[H_HASDISP] = parsed.apply(lambda p: "Yes" if p["raw"] is not None else "No")
        d[H_DISP_CLASS] = parsed.apply(lambda p: p["class"])
        d[H_DISP_PARTY] = parsed.apply(lambda p: p["party"])
        d[H_DISP_OUTCOME] = parsed.apply(lambda p: p["outcome"])
        d[H_DISP_CONN] = parsed.apply(
            lambda p: "Yes" if p["connected"] else ("No" if p["connected"] is False else ""))
    return d.drop(columns=["_est_dt"], errors="ignore")


def _q(v):
    """Quote a value as an Excel text criterion (exact match)."""
    return '"' + str(v).replace('"', '""') + '"'


class RawRef:
    """Addresses columns of an embedded raw sheet by header name and builds the
    aggregate-formula expressions (no leading '='; use EQ() to drop in a cell)."""

    def __init__(self, title, headers, nrows):
        self.title = title
        self.nrows = nrows
        self._col = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}

    def has(self, header):
        return header in self._col

    def rng(self, header):
        L = self._col[header]
        return f"'{self.title}'!${L}$2:${L}${self.nrows + 1}"

    def counta(self, header):
        return f"COUNTA({self.rng(header)})"

    def countif(self, header, crit):
        return f"COUNTIF({self.rng(header)},{crit})"

    def countifs(self, pairs):
        parts = []
        for h, c in pairs:
            parts += [self.rng(h), c]
        return "COUNTIFS(" + ",".join(parts) + ")"

    def sum(self, header):
        return f"SUM({self.rng(header)})"

    def sumif(self, crit_header, crit, sum_header):
        return f"SUMIF({self.rng(crit_header)},{crit},{self.rng(sum_header)})"

    def sumifs(self, sum_header, pairs):
        parts = [self.rng(sum_header)]
        for h, c in pairs:
            parts += [self.rng(h), c]
        return "SUMIFS(" + ",".join(parts) + ")"

    def averageifs(self, avg_header, pairs):
        parts = [self.rng(avg_header)]
        for h, c in pairs:
            parts += [self.rng(h), c]
        return "AVERAGEIFS(" + ",".join(parts) + ")"


def EQ(expr):
    return "=" + expr


def IFERR(expr, alt='"-"'):
    return f"IFERROR({expr},{alt})"


def write_raw_sheet(wb, title, df_out):
    """Dump the prepared raw frame to a sheet; return a RawRef for it."""
    ws = wb.create_sheet(title)
    headers = list(df_out.columns)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(1, j, str(h))
        c.font = f(9, True, "FFFFFF"); c.fill = fill(BLUE); c.border = BORDER
    for i, (_, row) in enumerate(df_out.iterrows(), start=2):
        for j, h in enumerate(headers, start=1):
            v = row[h]
            if v is None or (isinstance(v, float) and pd.isna(v)) or (
                    isinstance(v, type(pd.NaT)) and pd.isna(v)):
                v = None
            elif isinstance(v, float) and v != v:  # NaN guard
                v = None
            ws.cell(i, j, v)
    ws.freeze_panes = "A2"
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 15
    return RawRef(title, headers, len(df_out))


# --------------------------------------------------------------------------- #
# Workbook writer                                                              #
# --------------------------------------------------------------------------- #
class Sheet:
    def __init__(self, ws):
        self.ws = ws
        self.r = 1

    def title(self, text, sub=None):
        ws = self.ws
        c = ws.cell(self.r, 1, text)
        c.font = f(16, True, "FFFFFF"); c.fill = fill(NAVY)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[self.r].height = 26
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=8)
        self.r += 1
        if sub:
            c = ws.cell(self.r, 1, sub)
            c.font = f(9, False, "FFFFFF"); c.fill = fill(BLUE)
            c.alignment = Alignment(vertical="center", indent=1)
            ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=8)
            self.r += 1
        self.r += 1

    def section(self, text):
        c = self.ws.cell(self.r, 1, text)
        c.font = f(12, True, NAVY)
        self.r += 1

    def kv(self, key, val, valfmt=None, color="000000", bold_val=False):
        ws = self.ws
        a = ws.cell(self.r, 1, key); a.font = f(10, False, GREY)
        b = ws.cell(self.r, 2, val); b.font = f(11, bold_val, color)
        if valfmt:
            b.number_format = valfmt
        self.r += 1

    def note(self, text, color=GREY):
        c = self.ws.cell(self.r, 1, text); c.font = f(9, False, color)
        c.alignment = Alignment(wrap_text=False)
        self.r += 1

    def bullet(self, text, cols=8, color="000000", size=10):
        """A wrapped, full-width recommendation line with a bullet marker."""
        ws = self.ws
        c = ws.cell(self.r, 1, "•  " + text)
        c.font = f(size, False, color)
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.merge_cells(start_row=self.r, start_column=1,
                       end_row=self.r, end_column=cols)
        # rough auto-height: ~95 chars per line at this width
        lines = max(1, (len(text) // 95) + 1)
        ws.row_dimensions[self.r].height = 15 * lines + 3
        self.r += 1

    def kpi_row(self, items):
        """A horizontal strip of KPI cells: items = [(label, value, fmt, color)]."""
        ws = self.ws
        start = self.r
        for j, (label, value, fmt, color) in enumerate(items, start=1):
            lc = ws.cell(start, j, label)
            lc.font = f(8, False, "FFFFFF"); lc.fill = fill(BLUE)
            lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            lc.border = BORDER
            vc = ws.cell(start + 1, j, value)
            vc.font = f(14, True, color or "FFFFFF"); vc.fill = fill(NAVY)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = BORDER
            if fmt:
                vc.number_format = fmt
        ws.row_dimensions[start].height = 26
        ws.row_dimensions[start + 1].height = 24
        self.r += 2

    def blank(self, n=1):
        self.r += n

    def table(self, headers, rows, pct_cols=(), num_cols=(), money_cols=(),
              widths=None, total_row=False):
        ws = self.ws
        hr = self.r
        for j, h in enumerate(headers, start=1):
            c = ws.cell(hr, j, h)
            c.font = f(10, True, "FFFFFF"); c.fill = fill(BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        self.r += 1
        for i, row in enumerate(rows):
            is_tot = total_row and i == len(rows) - 1
            for j, val in enumerate(row, start=1):
                c = ws.cell(self.r, j, val)
                c.border = BORDER
                c.font = f(10, is_tot)
                if is_tot:
                    c.fill = fill(LIGHT)
                elif i % 2:
                    c.fill = fill(LIGHT2)
                if j in pct_cols:
                    c.number_format = "0.0%"
                    c.alignment = Alignment(horizontal="center")
                elif j in money_cols:
                    c.number_format = '$#,##0.000;($#,##0.000);-'
                    c.alignment = Alignment(horizontal="right")
                elif j in num_cols:
                    c.number_format = "#,##0"
                    c.alignment = Alignment(horizontal="center")
                elif j > 1:
                    c.alignment = Alignment(horizontal="center")
            self.r += 1
        if widths:
            for j, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(j)].width = w
        self.r += 1


def best_bucket(series):
    if series is None or series.empty:
        return ("-", 0, 0.0)
    tot = series.sum()
    b = series.idxmax()
    return (b, int(series.max()), pct(series.max(), tot))


def _count_pct_table(sh, headers, label_count_exprs, widths, *, total_label="Total"):
    """Write a 3-col 'label | count | % of total' table where count cells are
    live COUNTIF/COUNTIFS formulas and the % + Total cells are live cell-ratio /
    SUM formulas. label_count_exprs = [(label, count_expr_without_eq), ...]."""
    data_start = sh.r + 1
    total_row = data_start + len(label_count_exprs)
    rows = []
    for i, (label, cexpr) in enumerate(label_count_exprs):
        er = data_start + i
        rows.append([label, EQ(cexpr), EQ(IFERR(f"B{er}/B${total_row}", "0"))])
    rows.append([total_label, EQ(f"SUM(B{data_start}:B{total_row - 1})"),
                 EQ(IFERR(f"B{total_row}/B${total_row}", "0"))])
    sh.table(list(headers), rows, pct_cols=(3,), num_cols=(2,),
             widths=list(widths), total_row=True)


def write_distribution_table(sh, title, raw, bucket_header, present_buckets, value_label):
    """Missed-by-window distribution: count = COUNTIFS(bucket, Is Connect=No)."""
    sh.section(title)
    if not present_buckets:
        sh.table(["Time Bucket", value_label, "% Distribution"],
                 [["(no data)", 0, 0.0]], pct_cols=(3,), num_cols=(2,),
                 widths=[18, 16, 16])
        return
    exprs = [(b, raw.countifs([(bucket_header, _q(b)), (H_ISCONNECT, _q("No"))]))
             for b in present_buckets]
    _count_pct_table(sh, ["Time Bucket", value_label, "% Distribution"],
                     exprs, (18, 16, 16))


def write_outcome_matrix(sh, title, raw, bucket_header, result_header, ct):
    cols = [c for c in ct.columns if c != "Total"]
    ncols = len(cols)
    headers = ["Time Bucket"] + cols + ["Total"]
    last_res = get_column_letter(1 + ncols)        # results occupy cols 2..1+ncols
    buckets = list(ct.index)
    sh.section(title)
    data_start = sh.r + 1   # first data row (after section + table header)
    total_row = data_start + len(buckets)
    rows = []
    for i, b in enumerate(buckets):
        er = data_start + i
        rowvals = [b]
        for c in cols:
            rowvals.append(EQ(raw.countifs([(bucket_header, _q(b)), (result_header, _q(c))])))
        rowvals.append(EQ(f"SUM(B{er}:{last_res}{er})"))
        rows.append(rowvals)
    trow = ["Total"]
    for k in range(ncols):
        colL = get_column_letter(2 + k)
        trow.append(EQ(f"SUM({colL}{data_start}:{colL}{total_row - 1})"))
    trow.append(EQ(f"SUM(B{total_row}:{last_res}{total_row})"))
    rows.append(trow)
    widths = [18] + [max(10, len(str(c)) * 0.95) for c in cols] + [10]
    sh.table(headers, rows, num_cols=tuple(range(2, len(headers) + 1)),
             widths=widths, total_row=True)


def write_outcome_matrix_pct(sh, title, raw, bucket_header, result_header, ct):
    """Within-bucket % of each outcome: COUNTIFS / per-bucket Calls cell."""
    cols = [c for c in ct.columns if c != "Total"]
    ncols = len(cols)
    calls_col = get_column_letter(2 + ncols)        # the 'Calls' column
    headers = ["Time Bucket"] + cols + ["Calls"]
    buckets = list(ct.index)
    sh.section(title)
    data_start = sh.r + 1   # first data row (after section + table header)
    rows = []
    for i, b in enumerate(buckets):
        er = data_start + i
        rowvals = [b]
        for c in cols:
            num = raw.countifs([(bucket_header, _q(b)), (result_header, _q(c))])
            rowvals.append(EQ(IFERR(f"{num}/{calls_col}{er}", "0")))
        rowvals.append(EQ(raw.countif(bucket_header, _q(b))))
        rows.append(rowvals)
    widths = [18] + [max(10, len(str(c)) * 0.95) for c in cols] + [9]
    sh.table(headers, rows, pct_cols=tuple(range(2, ncols + 2)),
             num_cols=(len(headers),), widths=widths)


def write_rate_table(sh, title, raw, key_header, key_label, present_keys,
                     success_label, note=None):
    """Count + success + rate by weekday or week-block, fully formula-driven."""
    sh.section(title)
    if note:
        sh.note(note)
    data_start = sh.r + 1
    rows = []
    if present_keys:
        total_row = data_start + len(present_keys)
        for i, k in enumerate(present_keys):
            er = data_start + i
            rows.append([k, EQ(raw.countif(key_header, _q(k))),
                         EQ(raw.countifs([(key_header, _q(k)), (H_ISCONNECT, _q("Yes"))])),
                         EQ(IFERR(f"C{er}/B{er}", "0"))])
        rows.append(["Total", EQ(f"SUM(B{data_start}:B{total_row - 1})"),
                     EQ(f"SUM(C{data_start}:C{total_row - 1})"),
                     EQ(IFERR(f"C{total_row}/B{total_row}", "0"))])
    else:
        rows.append(["(no data)", 0, 0, 0.0])
    sh.table([key_label, "Calls", success_label, "Connect %"], rows,
             pct_cols=(4,), num_cols=(2, 3), widths=[20, 14, 18, 14],
             total_row=bool(present_keys))


def write_best_day_per_week(sh, title, raw, block_dow):
    """Monthly: per week-block, the best day (chosen in pandas) with live
    COUNTIFS numbers behind it."""
    sh.section(title)
    data_start = sh.r + 1
    rows = []
    for i, b in enumerate(WEEKBLOCK_ORDER):
        er = data_start + i
        frame = block_dow.get(b)
        bd = best_by_rate(frame) if frame is not None else None
        if bd:
            day = bd[0]
            rows.append([b, day,
                         EQ(raw.countifs([(H_WEEKBLOCK, _q(b)), (H_WEEKDAY, _q(day)),
                                          (H_ISCONNECT, _q("Yes"))])),
                         EQ(raw.countifs([(H_WEEKBLOCK, _q(b)), (H_WEEKDAY, _q(day))])),
                         EQ(IFERR(f"C{er}/D{er}", "0"))])
        else:
            rows.append([b, "(no data)", 0, 0, 0.0])
    sh.table(["Week", "Best Day", "Success", "Calls", "Connect %"], rows,
             pct_cols=(5,), num_cols=(3, 4), widths=[18, 14, 12, 12, 14])


def write_rate_window_table(sh, title, raw, bucket_header, present_buckets,
                            conn_label, cost_header=None, note=None):
    """Connect-RATE by time window, all live formulas: dials = COUNTIF(bucket),
    connects = COUNTIFS(bucket, Is Connect=Yes), rate/dials-per-connect = cell
    ratios, $/connect = SUMIF(spend in window) / connects."""
    sh.section(title)
    if note:
        sh.note(note)
    has_cost = bool(cost_header) and raw.has(cost_header)
    headers = ["Time Window", "Dials", conn_label, "Connect %", "Dials / Connect"]
    if has_cost:
        headers.append("$ / Connect")
    data_start = sh.r + 1
    rows = []
    if present_buckets:
        total_row = data_start + len(present_buckets)
        for i, b in enumerate(present_buckets):
            er = data_start + i
            row = [b, EQ(raw.countif(bucket_header, _q(b))),
                   EQ(raw.countifs([(bucket_header, _q(b)), (H_ISCONNECT, _q("Yes"))])),
                   EQ(IFERR(f"C{er}/B{er}", "0")),
                   EQ(IFERR(f"B{er}/C{er}", '"-"'))]
            if has_cost:
                spend = raw.sumif(bucket_header, _q(b), cost_header)
                row.append(EQ(IFERR(f"{spend}/C{er}", '"-"')))
            rows.append(row)
        trow = ["Total", EQ(f"SUM(B{data_start}:B{total_row - 1})"),
                EQ(f"SUM(C{data_start}:C{total_row - 1})"),
                EQ(IFERR(f"C{total_row}/B{total_row}", "0")),
                EQ(IFERR(f"B{total_row}/C{total_row}", '"-"'))]
        if has_cost:
            trow.append("")
        rows.append(trow)
    else:
        rows.append(["(no data)", 0, 0, 0.0, "-"] + (["-"] if has_cost else []))
    widths = [20, 12, 14, 12, 16] + ([14] if has_cost else [])
    money = (6,) if has_cost else ()
    sh.table(headers, rows, pct_cols=(4,), num_cols=(2, 3), money_cols=money,
             widths=widths, total_row=bool(present_buckets))


def write_heatmap(sh, title, raw, bucket_header, rate_mat, tot_mat, note=None):
    """Weekday x time-window grid. Rate cells = COUNTIFS(connected)/COUNTIFS(all)
    live; volume cells = COUNTIFS(all) live. Red->green color scale on rates."""
    sh.section(title)
    if note:
        sh.note(note)
    if rate_mat is None or rate_mat.empty:
        sh.note("(needs multiple days of data)")
        return
    ws = sh.ws
    cols = list(rate_mat.columns)
    # ---- rate grid ----
    hr = sh.r
    ws.cell(hr, 1, "Connect %").font = f(10, True, "FFFFFF")
    ws.cell(hr, 1).fill = fill(BLUE); ws.cell(hr, 1).border = BORDER
    ws.cell(hr, 1).alignment = Alignment(horizontal="center", vertical="center")
    for j, c in enumerate(cols, start=2):
        cell = ws.cell(hr, j, c)
        cell.font = f(8, True, "FFFFFF"); cell.fill = fill(BLUE); cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sh.r += 1
    first_data = sh.r
    for day in rate_mat.index:
        cell = ws.cell(sh.r, 1, day)
        cell.font = f(9, True); cell.border = BORDER
        for j, c in enumerate(cols, start=2):
            num = raw.countifs([(H_WEEKDAY, _q(day)), (bucket_header, _q(c)),
                                (H_ISCONNECT, _q("Yes"))])
            den = raw.countifs([(H_WEEKDAY, _q(day)), (bucket_header, _q(c))])
            cc = ws.cell(sh.r, j, EQ(IFERR(f"{num}/{den}", '""')))
            cc.border = BORDER
            cc.alignment = Alignment(horizontal="center")
            cc.number_format = "0%"
        sh.r += 1
    last_data = sh.r - 1
    rng = f"{get_column_letter(2)}{first_data}:{get_column_letter(1 + len(cols))}{last_data}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"))
    ws.column_dimensions["A"].width = 14
    for j in range(2, 2 + len(cols)):
        ws.column_dimensions[get_column_letter(j)].width = 9
    sh.r += 1
    # ---- volume grid ----
    sh.section("Dial volume by the same windows (context for the rates above)")
    hr = sh.r
    ws.cell(hr, 1, "Dials").font = f(10, True, "FFFFFF")
    ws.cell(hr, 1).fill = fill(BLUE); ws.cell(hr, 1).border = BORDER
    for j, c in enumerate(cols, start=2):
        cell = ws.cell(hr, j, c)
        cell.font = f(8, True, "FFFFFF"); cell.fill = fill(BLUE); cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sh.r += 1
    for day in tot_mat.index:
        ws.cell(sh.r, 1, day).font = f(9, True); ws.cell(sh.r, 1).border = BORDER
        for j, c in enumerate(cols, start=2):
            cc = ws.cell(sh.r, j, EQ(raw.countifs([(H_WEEKDAY, _q(day)),
                                                   (bucket_header, _q(c))])))
            cc.border = BORDER; cc.number_format = "#,##0"
            cc.alignment = Alignment(horizontal="center")
        sh.r += 1
    sh.r += 1


def write_recommendations(sh, recs):
    sh.section("Recommended actions")
    for line in recs:
        sh.bullet(line)
    sh.blank()


# --------------------------------------------------------------------------- #
# Trend writers (WoW / MoM / DoD)  --  raw deltas, ▲/▼ direction, color = good  #
# --------------------------------------------------------------------------- #
def _move_color(good):
    return GREEN if good else (RED if good is False else GREY)


def _write_trend_table(sh, stream_label, rows):
    """Metric | Prior | Current | Δ | %Δ, with the Δ/%Δ cells colored by whether
    the move is good (green) or bad (red) for that metric."""
    if not rows:
        return
    ws = sh.ws
    sh.section(stream_label)
    headers = ["Metric", "Prior", "Current", "Δ", "% Δ"]
    hr = sh.r
    for j, h in enumerate(headers, start=1):
        c = ws.cell(hr, j, h)
        c.font = f(10, True, "FFFFFF"); c.fill = fill(BLUE); c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    sh.r += 1
    for i, row in enumerate(rows):
        rr = sh.r
        shade = fill(LIGHT2) if i % 2 else None
        color = _move_color(row["good"])
        cells = [
            (1, row["metric"], None, f(10), "left"),
            (2, row["prior"], row["fmt"], f(10), "center"),
            (3, row["current"], row["fmt"], f(10), "center"),
            (4, f'{row["arrow"]} {row["delta_str"]}', None, f(10, True, color), "center"),
            (5, row["pct_str"], None, f(10, False, color), "center"),
        ]
        for col, val, numfmt, font, align in cells:
            c = ws.cell(rr, col, val)
            c.border = BORDER; c.font = font
            c.alignment = Alignment(horizontal=align)
            if numfmt:
                c.number_format = numfmt
            if shade:
                c.fill = shade
        sh.r += 1
    ws.column_dimensions["A"].width = 20
    for col in "BCDE":
        ws.column_dimensions[col].width = 14
    sh.r += 1


def _disp_delta_str(d):
    sign = "+" if d > 1e-9 else ("−" if d < -1e-9 else "")
    arrow = "▲" if d > 1e-9 else ("▼" if d < -1e-9 else "▬")
    return f"{arrow} {sign}{abs(d) * 100:.1f} pts"


def _write_disp_trend(sh, moves, label="Inbound"):
    """Disposition-mix movement table (share of dispositioned calls)."""
    if not moves:
        return
    sh.section(f"{label} — disposition mix movement (share of dispositioned calls)")
    rows = []
    for m in moves[:12]:
        tag = "  (new)" if m["is_new"] else ("  (gone)" if m["is_gone"] else "")
        rows.append([f"{m['code']}{tag}", m["prior_share"], m["current_share"],
                     _disp_delta_str(m["share_delta"]),
                     m["prior_count"], m["current_count"]])
    sh.table(["Disposition", "Prior %", "Current %", "Δ", "Prior calls", "Current calls"],
             rows, pct_cols=(2, 3), num_cols=(5, 6),
             widths=[26, 12, 12, 14, 12, 13])


def _write_agent_trend(sh, moves, label="Inbound"):
    """Per-agent connect-rate movement table (who went up / down)."""
    if not moves:
        return
    elig = [m for m in moves if m["rate_delta"] is not None]
    if not elig:
        return
    sh.section(f"{label} — agent connect-rate movement")
    rows = []
    for m in elig[:15]:
        rows.append([m["agent"], m["prior_rate"], m["current_rate"],
                     _disp_delta_str(m["rate_delta"]),
                     m["prior_handled"], m["current_handled"]])
    sh.table(["Agent", "Prior connect %", "Current connect %", "Δ",
              "Prior calls", "Current calls"],
             rows, pct_cols=(2, 3), num_cols=(5, 6),
             widths=[22, 15, 16, 14, 12, 13])


def write_trends_tab(wb, client, comparisons):
    """Standalone Trends tab: one section per comparison basis (WoW/MoM/DoD)."""
    ws = wb.create_sheet("Trends"); ws.sheet_view.showGridLines = False
    sh = Sheet(ws)
    sh.title(f"{client} — Trends",
             "This period vs the prior period(s). Raw deltas — ▲/▼ = direction, "
             "green = better, red = worse. Δ on rates is in percentage points.")
    for blk in comparisons:
        sh.section(f"{blk['basis']}   ({blk['prior_label']}  →  {blk['current_label']})")
        _write_trend_table(sh, "Outbound", blk["ob"])
        _write_trend_table(sh, "Inbound", blk["ib"])
        if blk["best_window_shift"]:
            pw, cw = blk["best_window_shift"]
            sh.kv("Best calling window shifted", f"{pw}  →  {cw}",
                  color=AMBER, bold_val=True)
        _write_disp_trend(sh, blk.get("ob_disp"), label="Outbound")
        _write_agent_trend(sh, blk.get("ob_agents"), label="Outbound")
        _write_disp_trend(sh, blk.get("ib_disp"), label="Inbound")
        _write_agent_trend(sh, blk.get("ib_agents"), label="Inbound")
        sh.blank()


def write_leadership_movement(sh, comparisons):
    """Compact 'what moved' block for the Leadership View hub page."""
    if not comparisons:
        sh.section("Movement vs prior period")
        sh.note("No prior period of this type yet — this run is the baseline. "
                "Trends appear automatically from the next comparable period.")
        sh.blank()
        return
    blk = comparisons[0]
    sh.section(f"Movement — {blk['basis']} (vs {blk['prior_label']})")
    for r in blk["ob"][:4]:
        sh.kv(f"OB {r['metric']}",
              f"{r['arrow']} {r['delta_str']}  ({r['pct_str']})",
              color=_move_color(r["good"]), bold_val=True)
    for r in blk["ib"][:2]:
        sh.kv(f"IB {r['metric']}",
              f"{r['arrow']} {r['delta_str']}  ({r['pct_str']})",
              color=_move_color(r["good"]), bold_val=True)
    if blk["best_window_shift"]:
        pw, cw = blk["best_window_shift"]
        sh.kv("Best window shifted", f"{pw} → {cw}", color=AMBER)
    if len(comparisons) > 1:
        sh.note(f"Also compared: {comparisons[1]['basis']} — full detail on the Trends tab.")
    sh.blank()


def write_ib_disposition_tabs(wb, ad, raw, talk_h=None, prefix="IB", label="Inbound"):
    """Two current-period drill-down tabs from the agent-disposition layer:
    '<prefix> Agents' (per-agent performance) and '<prefix> Dispositions'
    (response-code distribution + class/party/outcome rollups). All numbers are
    live COUNTIFS/AVERAGEIFS formulas over the embedded <prefix>_RawData sheet.
    No-op when the dump has no disposition column. `talk_h` = the dump's
    talk-duration column name (for live AVERAGEIFS), or None to fall back to the
    computed avg. Used for both inbound and outbound."""
    if not ad:
        return
    has_talk = bool(talk_h) and raw.has(talk_h)
    # --------------------------- <prefix> Agents --------------------------- #
    ws = wb.create_sheet(f"{prefix} Agents"); ws.sheet_view.showGridLines = False
    sh = Sheet(ws)
    sh.title(f"{label} — Agent Performance",
             "Per-agent outcomes on agent-handled calls. Connect = response "
             "code starts C-. Ranked by calls handled.")
    agents = ad["agents"]
    if agents:
        sh.section(f"Agents ranked by calls handled ({ad['n_agents']} agent(s))")
        data_start = sh.r + 1
        rows = []
        for i, r in enumerate(agents):
            er = data_start + i
            a = _q(r["agent"])
            handled = raw.countifs([(H_AGENT, a), (H_HASDISP, _q("Yes"))])
            conn = raw.countifs([(H_AGENT, a), (H_HASDISP, _q("Yes")),
                                 (H_DISP_CONN, _q("Yes"))])
            ptp = raw.countifs([(H_AGENT, a), (H_HASDISP, _q("Yes")),
                                (H_DISP_OUTCOME, _q("PTP"))])
            pay = raw.countifs([(H_AGENT, a), (H_HASDISP, _q("Yes")),
                                (H_DISP_OUTCOME, _q("PAYMENT"))])
            row = [r["agent"], EQ(handled), EQ(conn),
                   EQ(IFERR(f"C{er}/B{er}", "0")), EQ(ptp), EQ(pay)]
            # avg talk = AVERAGEIFS over the dump's talk column if present
            if has_talk:
                row.append(EQ(IFERR(raw.averageifs(
                    talk_h, [(H_AGENT, a), (H_HASDISP, _q("Yes"))]), "0")))
            else:
                row.append(round(r["avg_talk"], 1))
            row.append(r["top_disp"])
            rows.append(row)
        sh.table(["Agent", "Handled", "Connected", "Connect %", "PTP",
                  "Payment", "Avg talk (s)", "Top disposition"], rows,
                 pct_cols=(4,), num_cols=(2, 3, 5, 6, 7),
                 widths=[22, 10, 11, 11, 8, 9, 13, 24])
    else:
        sh.section("No named agents this period")
        sh.note("Every inbound call this period was system-handled "
                "(abandoned / voicemail / not connected).")
    if ad.get("agent_blank"):
        sh.blank()
        sh.section("System / Unassigned (not agent-handled)")
        sh.kv("Calls (abandoned / voicemail / not connected)",
              EQ(raw.countif(H_AGENT, _q(AGENT_BLANK_LABEL))),
              "#,##0", color=GREY, bold_val=True)

    # ----------------------- <prefix> Dispositions ----------------------- #
    ws = wb.create_sheet(f"{prefix} Dispositions"); ws.sheet_view.showGridLines = False
    sh = Sheet(ws)
    sh.title(f"{label} — Agent Dispositions",
             "'An Agent Call Response' breakdown. Codes parse as "
             "class (C/NC) · party (RPC/TPC) · outcome.")
    sh.section("Connectivity (of dispositioned calls)")
    disp_total_expr = raw.countif(H_HASDISP, _q("Yes"))
    disp_total_row = sh.r
    sh.kv("Dispositioned calls", EQ(disp_total_expr), "#,##0", bold_val=True)
    sh.kv("Connected (C-*)", EQ(raw.countif(H_DISP_CONN, _q("Yes"))),
          "#,##0", color=GREEN, bold_val=True)
    sh.kv("Connected rate", EQ(IFERR(f"B{disp_total_row + 1}/B{disp_total_row}", "0")),
          "0.0%", bold_val=True)
    sh.blank()
    denom = f"$B${disp_total_row}"
    dc = ad["disp_counts"]
    sh.section("Full disposition distribution")
    _count_pct_table(sh, ["Disposition", "Calls", "% of dispositioned"],
                     [(idx, raw.countif(H_DISP_CODE, _q(idx))) for idx in dc.index],
                     (28, 14, 18))
    for title2, series, crit_h in (
            ("By contact class", ad.get("by_class"), H_DISP_CLASS),
            ("By party", ad.get("by_party"), H_DISP_PARTY),
            ("By outcome", ad.get("by_outcome"), H_DISP_OUTCOME)):
        if series is not None and len(series):
            sh.section(title2)
            data_start = sh.r + 1
            rows = []
            for i, idx in enumerate(series.index):
                er = data_start + i
                rows.append([idx, EQ(raw.countif(crit_h, _q(idx))),
                             EQ(IFERR(f"B{er}/{denom}", "0"))])
            sh.table(["Segment", "Calls", "% of dispositioned"], rows,
                     pct_cols=(3,), num_cols=(2,), widths=[28, 14, 18])


# --------------------------------------------------------------------------- #
# Alert writers  --  flagged changes (severity), beside the detailed Trends    #
# --------------------------------------------------------------------------- #
SEV_FILL = {"CRITICAL": RED, "WARN": AMBER, "INFO": BLUE}
SEV_TEXT = {"CRITICAL": RED, "WARN": AMBER, "INFO": GREY}


def write_alerts_tab(wb, client, alerts):
    """Standalone Alerts tab: every fired alert with severity, the metric, the
    basis/window it was measured against, prior/current values, the change, and
    a plain-English reason."""
    ws = wb.create_sheet("Alerts"); ws.sheet_view.showGridLines = False
    sh = Sheet(ws)
    sh.title(f"{client} — Alerts",
             "Flagged changes worth attention this period. 🔴 critical · 🟠 warning "
             "· 🔵 info. Full unfiltered movement is on the Trends tab.")
    if not alerts:
        sh.section("No alerts")
        sh.note("No thresholds were tripped this period — every metric is within "
                "its configured normal range. (Tune thresholds in config.py → ALERTS.)")
        ws.column_dimensions["A"].width = 36
        return

    n_crit = sum(1 for a in alerts if a["severity"] == "CRITICAL")
    n_warn = sum(1 for a in alerts if a["severity"] == "WARN")
    n_info = sum(1 for a in alerts if a["severity"] == "INFO")
    sh.kv("Summary", f"{n_crit} critical · {n_warn} warning · {n_info} info",
          color=(RED if n_crit else (AMBER if n_warn else GREY)), bold_val=True)
    sh.blank()

    headers = ["Severity", "Stream", "Metric", "Basis / Window",
               "Prior", "Current", "Change", "What it means"]
    hr = sh.r
    for j, h in enumerate(headers, start=1):
        c = ws.cell(hr, j, h)
        c.font = f(10, True, "FFFFFF"); c.fill = fill(BLUE); c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sh.r += 1
    for i, a in enumerate(alerts):
        rr = sh.r
        shade = fill(LIGHT2) if i % 2 else None
        cells = [
            (1, a["severity"], f(9, True, "FFFFFF"), "center", fill(SEV_FILL.get(a["severity"], GREY))),
            (2, a["stream"], f(10), "center", shade),
            (3, a["metric"], f(10), "left", shade),
            (4, a["basis"] if a.get("window", "—") == "—" else f'{a["basis"]} · {a["window"]}',
                f(10), "left", shade),
            (5, a.get("prior_str", "—"), f(10), "center", shade),
            (6, a.get("current_str", "—"), f(10), "center", shade),
            (7, a.get("change_str", "—"), f(10, True, SEV_TEXT.get(a["severity"], GREY)), "center", shade),
            (8, a["message"], f(9), "left", shade),
        ]
        for col, val, font, align, cellfill in cells:
            c = ws.cell(rr, col, val)
            c.border = BORDER; c.font = font
            c.alignment = Alignment(horizontal=align, vertical="center",
                                    wrap_text=(col == 8))
            if cellfill:
                c.fill = cellfill
        sh.r += 1
    for col, w in zip("ABCDEFGH", [10, 9, 16, 18, 11, 11, 12, 52]):
        ws.column_dimensions[col].width = w


def write_leadership_alerts(sh, alerts):
    """Compact 'alerts this run' block for the Leadership View hub page."""
    sh.section("Alerts this run")
    if not alerts:
        sh.note("No threshold alerts — every metric is within its normal range.")
        sh.blank()
        return
    for a in alerts[:6]:
        marker = {"CRITICAL": "🔴", "WARN": "🟠", "INFO": "🔵"}.get(a["severity"], "•")
        sh.kv(f"{marker} {a['stream']} · {a['metric']}", a["message"],
              color=SEV_TEXT.get(a["severity"], GREY), bold_val=(a["severity"] == "CRITICAL"))
    if len(alerts) > 6:
        sh.note(f"+ {len(alerts) - 6} more — see the Alerts tab.")
    else:
        sh.note("Full detail on the Alerts tab.")
    sh.blank()


def _stream_period_tabs(wb, prefix, res, period, success_label, raw):
    """Add period-specific tabs (Day of Week for weekly+monthly; Week Breakdown
    + Best Day per Week for monthly) for one stream (OB or IB). All numbers are
    live COUNTIF/COUNTIFS formulas over `raw` (the embedded *_RawData sheet)."""
    if period in ("weekly", "monthly"):
        ws = wb.create_sheet(f"{prefix} Day of Week"); ws.sheet_view.showGridLines = False
        sh = Sheet(ws)
        sh.title(f"{prefix} — Day of Week",
                 "Connectivity by weekday. 'Best' day = highest connect rate.")
        dow = res.get("dow")
        write_rate_table(sh, "By day of week", raw, H_WEEKDAY, "Day",
                         list(dow.index) if dow is not None else [], success_label)
        bd = best_by_rate(res.get("dow"))
        wd = best_by_rate(res.get("dow"), want="worst")
        if bd:
            sh.kv("Best day (by rate)", f"{bd[0]} — {bd[3]:.1%} ({bd[1]}/{bd[2]})",
                  color=GREEN, bold_val=True)
        if wd:
            sh.kv("Worst day (by rate)", f"{wd[0]} — {wd[3]:.1%} ({wd[1]}/{wd[2]})",
                  color=RED, bold_val=True)

    if period == "monthly":
        ws = wb.create_sheet(f"{prefix} Week Breakdown"); ws.sheet_view.showGridLines = False
        sh = Sheet(ws)
        sh.title(f"{prefix} — Week Breakdown",
                 "Weeks: 1-7, 8-14, 15-21, 22-end. 'Best' week = highest connect rate.")
        wkb = res.get("weekblock")
        write_rate_table(sh, "By week of month", raw, H_WEEKBLOCK, "Week",
                         list(wkb.index) if wkb is not None else [], success_label)
        bw = best_by_rate(res.get("weekblock"))
        ww = best_by_rate(res.get("weekblock"), want="worst")
        if bw:
            sh.kv("Best week (by rate)", f"{bw[0]} — {bw[3]:.1%} ({bw[1]}/{bw[2]})",
                  color=GREEN, bold_val=True)
        if ww:
            sh.kv("Worst week (by rate)", f"{ww[0]} — {ww[3]:.1%} ({ww[1]}/{ww[2]})",
                  color=RED, bold_val=True)
        sh.blank()
        write_best_day_per_week(sh, "Best day within each week", raw,
                                res.get("block_dow", {}))


def build_workbook(client, period, date_label, ob_res, ib_res, out_path,
                   comparisons=None, alerts=None):
    wb = Workbook()
    # ---- prepare embedded raw-data sheets (every analytical formula points
    # here so leadership can click a cell and see the COUNTIF/SUMIF derivation).
    # The sheets themselves are written at the very end; we only need their
    # RawRef (title + headers + row count) to author the formulas now. ----
    ob_meta = ob_res["_df"].attrs["meta"] if ob_res else None
    ib_meta = ib_res["_df"].attrs["meta"] if ib_res else None
    ob_raw_df = _prep_raw(ob_res["_df"], ob_meta, "OB") if ob_res else None
    ib_raw_df = _prep_raw(ib_res["_df"], ib_meta, "IB") if ib_res else None
    ob_raw = RawRef("OB_RawData", list(ob_raw_df.columns), len(ob_raw_df)) if ob_res else None
    ib_raw = RawRef("IB_RawData", list(ib_raw_df.columns), len(ib_raw_df)) if ib_res else None
    ob_result_h = ob_meta["result"] if ob_res else None
    ob_cost_h = ob_meta.get("total_cost") if ob_res else None
    ib_result_h = ib_meta["result"] if ib_res else None
    # ===================================================================== #
    # LEADERSHIP VIEW  --  the one page leaders read; everything else is a   #
    # drill-down ("double click") tab.                                       #
    # ===================================================================== #
    ws = wb.active; ws.title = "Leadership View"
    ws.sheet_view.showGridLines = False
    sh = Sheet(ws)
    sh.title(f"{client} — Leadership View",
             f"{period.title()} · {date_label} · how to win more connects · "
             f"generated {datetime.now():%Y-%m-%d %H:%M}")

    if ob_res:
        sh.section("Outbound — the bottom line")
        vr = sh.r + 1   # KPI value row (formulas reference cells in this row)
        items = [("Dials", EQ(ob_raw.counta(ob_result_h)), "#,##0", "FFFFFF"),
                 ("Connects", EQ(ob_raw.countif(H_ISCONNECT, _q("Yes"))), "#,##0", "C6E0B4"),
                 ("Connect rate", EQ(IFERR(f"B{vr}/A{vr}", "0")), "0.0%", "C6E0B4"),
                 ("Dials / connect", EQ(IFERR(f"A{vr}/B{vr}", '"-"')), "0.0", "FFFFFF")]
        if ob_res.get("total_cost") is not None and ob_cost_h:
            spend = ob_raw.sum(ob_cost_h)
            items.append(("$ / connect", EQ(IFERR(f"({spend})/B{vr}", "0")), '$#,##0.000', "FFE699"))
            items.append(("Total spend", EQ(spend), '$#,##0', "FFFFFF"))
        sh.kpi_row(items)
        sh.blank()

        sh.section("When to call — highest connect-rate windows (EST)")
        peaks = peak_windows(ob_res.get("rate_est"))
        if peaks:
            rows = [[w, t, s, r] for (w, s, t, r) in peaks]
            sh.table(["Top windows", "Dials", "Connects", "Connect %"], rows,
                     pct_cols=(4,), num_cols=(2, 3), widths=[20, 12, 12, 12])
        if period in ("weekly", "monthly"):
            bd = best_by_rate(ob_res.get("dow"))
            if bd:
                sh.kv("Best day", f"{bd[0]} — {bd[3]:.1%} ({bd[1]}/{bd[2]})",
                      color=GREEN, bold_val=True)
        if period == "monthly":
            bw = best_by_rate(ob_res.get("weekblock"))
            if bw:
                sh.kv("Best week", f"{bw[0]} — {bw[3]:.1%} ({bw[1]}/{bw[2]})",
                      color=GREEN, bold_val=True)
        sh.blank()

        sh.section("Where we lose connects")
        worst = best_by_rate(ob_res.get("rate_est"), want="worst")
        if worst:
            sh.kv("Weakest window (EST)",
                  f"{worst[0]} — {worst[3]:.1%} ({worst[1]}/{worst[2]})", color=RED)
        wc = ob_res.get("waste_counts")
        if wc is not None and len(wc) and ob_res["total"]:
            for idx, v in wc.head(3).items():
                sh.kv(f"   {idx}", f"{int(v):,} dials  ({v / ob_res['total']:.0%} of dials)")
        sh.blank()

        if ob_res.get("total_cost") is not None:
            sh.section("Cost efficiency")
            sh.kv("Cost per connect", ob_res.get("cost_per_al", 0), '$#,##0.000',
                  color=AMBER, bold_val=True)
            cpc = ob_res.get("cost_per_connect_est")
            if cpc is not None and len(cpc):
                sh.kv("Cheapest connect window",
                      f"{cpc.idxmin()} — ${cpc.min():,.3f}/connect", color=GREEN)
            if ob_res.get("cost_wasted") is not None:
                sh.kv("Spend on dials that didn't connect",
                      ob_res["cost_wasted"], '$#,##0.000', color=RED)
            sh.blank()

        oad = ob_res.get("agentdisp")
        if oad and (oad.get("agents") or len(oad.get("disp_counts", []))):
            sh.section("Outbound — agents & dispositions")
            sh.kv("Connected (C-*) rate of dispositioned dials",
                  EQ(IFERR(f"{ob_raw.countif(H_DISP_CONN, _q('Yes'))}/"
                           f"{ob_raw.countif(H_HASDISP, _q('Yes'))}", "0")),
                  "0.0%", color=GREEN, bold_val=True)
            if oad["agents"]:
                ranked = [a for a in oad["agents"] if a["handled"] >= 5] or oad["agents"]
                top = max(ranked, key=lambda a: a["rate"])
                low = min(ranked, key=lambda a: a["rate"])
                sh.kv("Top agent (connect rate, ≥5 dials)",
                      f"{top['agent']} — {top['rate']:.1%} ({top['handled']:,} dials)",
                      color=GREEN)
                if low["agent"] != top["agent"]:
                    sh.kv("Lowest agent (connect rate, ≥5 dials)",
                          f"{low['agent']} — {low['rate']:.1%} ({low['handled']:,} dials)",
                          color=RED)
                sh.kv("Named agents handling dials", oad["n_agents"], "#,##0")
            if len(oad["disp_counts"]):
                tc = oad["disp_counts"]
                sh.kv("Most common disposition",
                      f"{tc.index[0]} — {int(tc.iloc[0]):,} "
                      f"({tc.iloc[0] / oad['disp_total']:.0%})")
            sh.blank()

    if ib_res:
        sh.section("Inbound — the bottom line")
        vr = sh.r + 1
        items = [("Calls", EQ(ib_raw.counta(ib_result_h)), "#,##0", "FFFFFF"),
                 ("Connected", EQ(ib_raw.countif(H_ISCONNECT, _q("Yes"))), "#,##0", "C6E0B4"),
                 ("Connect rate", EQ(IFERR(f"B{vr}/A{vr}", "0")), "0.0%", "C6E0B4"),
                 ("Missed", EQ(ib_raw.countif(H_ISCONNECT, _q("No"))), "#,##0", "F4B0A6"),
                 ("Miss rate", EQ(IFERR(f"D{vr}/A{vr}", "0")), "0.0%", "F4B0A6")]
        sh.kpi_row(items)
        sh.blank()
        sh.section("Inbound staffing signal")
        mb = best_bucket(ib_res["miss_est"])
        sh.kv("Most-missed window (EST)", f"{mb[0]} — {mb[1]} missed calls", color=RED)
        be, ne, pe = best_bucket(ib_res["est_conn"])
        sh.kv("Busiest connected window (EST)", f"{be} — {ne} calls")
        if period in ("weekly", "monthly"):
            wd = best_by_rate(ib_res.get("dow"), want="worst")
            if wd:
                sh.kv("Worst connect day", f"{wd[0]} — {wd[3]:.1%}", color=RED)
        sh.blank()

        ad = ib_res.get("agentdisp")
        if ad:
            sh.section("Inbound — agents & dispositions")
            sh.kv("Connected (C-*) rate of dispositioned calls",
                  EQ(IFERR(f"{ib_raw.countif(H_DISP_CONN, _q('Yes'))}/"
                           f"{ib_raw.countif(H_HASDISP, _q('Yes'))}", "0")),
                  "0.0%", color=GREEN, bold_val=True)
            if ad["agents"]:
                ranked = [a for a in ad["agents"] if a["handled"] >= 5] or ad["agents"]
                top = max(ranked, key=lambda a: a["rate"])
                low = min(ranked, key=lambda a: a["rate"])
                sh.kv("Top agent (connect rate, ≥5 calls)",
                      f"{top['agent']} — {top['rate']:.1%} ({top['handled']:,} calls)",
                      color=GREEN)
                if low["agent"] != top["agent"]:
                    sh.kv("Lowest agent (connect rate, ≥5 calls)",
                          f"{low['agent']} — {low['rate']:.1%} ({low['handled']:,} calls)",
                          color=RED)
                sh.kv("Named agents handling calls", ad["n_agents"], "#,##0")
            if len(ad["disp_counts"]):
                tc = ad["disp_counts"]
                sh.kv("Most common disposition",
                      f"{tc.index[0]} — {int(tc.iloc[0]):,} "
                      f"({tc.iloc[0] / ad['disp_total']:.0%})")
            sh.blank()

    if alerts is not None:
        write_leadership_alerts(sh, alerts)

    write_leadership_movement(sh, comparisons)

    write_recommendations(sh, recommendations(period, ob_res, ib_res))

    sh.section("Where to look next — drill-down tabs")
    if alerts is not None:
        sh.note("• Alerts — flagged changes with severity (🔴/🟠/🔵), values, and why")
    if comparisons:
        sh.note("• Trends — full WoW/MoM/DoD movement for every metric (▲/▼, Δ, %Δ)")
    if ob_res:
        sh.note("• OB Best Time to Call — connect RATE (not just volume) by EST & IST window, with $/connect")
        if period in ("weekly", "monthly"):
            sh.note("• OB Heatmap — weekday × time-of-day connect-rate grid (your scheduling map)")
        sh.note("• OB Outcomes & Waste — every non-connect outcome by window")
        sh.note("• OB Cost — cost per connect, cheapest windows, connected vs wasted spend")
        if ob_res.get("agentdisp") and (ob_res["agentdisp"].get("agents")
                                        or len(ob_res["agentdisp"].get("disp_counts", []))):
            sh.note("• OB Agents — per-agent connect rate, PTP/payment, talk time (ranked)")
            sh.note("• OB Dispositions — agent-response breakdown by class / party / outcome")
        if period == "weekly":
            sh.note("• OB Day of Week")
        if period == "monthly":
            sh.note("• OB Day of Week  ·  OB Week Breakdown")
    if ib_res:
        sh.note("• IB Best Time & Missed — connect & miss rate by window (staffing guide)")
        if period in ("weekly", "monthly"):
            sh.note("• IB Heatmap — weekday × time-of-day connect-rate grid")
        sh.note("• IB Cost")
        if ib_res.get("agentdisp"):
            sh.note("• IB Agents — per-agent connect rate, PTP/payment, talk time (ranked)")
            sh.note("• IB Dispositions — agent-response breakdown by class / party / outcome")
    sh.blank()
    sh.section("Definitions")
    sh.note("• OB connect = Answered Linkcall (routed to an agent AND picked up). Connect rate = connects ÷ dials.")
    sh.note("• IB connect = Answered Linkcall + Answered. Missed = inbound that never reached an agent.")
    sh.note("• 'Connect-rate' windows rank by SUCCESS RATE (with a volume floor) — they show where to call, not just where calls happened.")
    sh.note("• Times in EST (campaign time) and IST (team time = EST + 10:30).")
    ws.column_dimensions["A"].width = 36
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 15

    # ----- Alerts tab (always when alerts were evaluated) ----- #
    if alerts is not None:
        write_alerts_tab(wb, client, alerts)

    # ----- Trends tab (only when we have a prior period to compare to) ----- #
    if comparisons:
        write_trends_tab(wb, client, comparisons)

    # ----- OB detail tabs ----- #
    if ob_res:
        ws = wb.create_sheet("OB Summary"); ws.sheet_view.showGridLines = False
        sh = Sheet(ws); sh.title("Outbound — Summary & Result Breakdown")
        rc = ob_res["result_counts"]; tot = ob_res["total"]
        sh.section("Result breakdown (all dials)")
        _count_pct_table(sh, ["Result", "Calls", "% of dials"],
                         [(idx, ob_raw.countif(ob_result_h, _q(idx))) for idx in rc.index],
                         (26, 14, 14))
        sh.section("Connectivity")
        total_dials = ob_raw.counta(ob_result_h)
        al_row = sh.r
        sh.kv("Answered linkcalls", EQ(ob_raw.countif(H_ISCONNECT, _q("Yes"))),
              "#,##0", color=GREEN, bold_val=True)
        sh.kv("Connectivity rate", EQ(IFERR(f"B{al_row}/({total_dials})", "0")),
              "0.0%", bold_val=True)
        sh.kv("Dials per connect", EQ(IFERR(f"({total_dials})/B{al_row}", '"-"')), "0.0")

        ws = wb.create_sheet("OB Best Time to Call"); ws.sheet_view.showGridLines = False
        sh = Sheet(ws); sh.title("Outbound — Best Time to Call",
                                 "Ranked by CONNECT RATE with volume + $/connect — your scheduling guide.")
        write_rate_window_table(sh, "Connect rate by EST window (campaign time)", ob_raw,
                                "EST Bucket", list(ob_res["rate_est"].index), "Connects",
                                cost_header=ob_cost_h,
                                note="Higher Connect % = better window. Dials/Connect = dials needed to expect one connect.")
        write_rate_window_table(sh, "Connect rate by IST window (team time)", ob_raw,
                                "IST Bucket", list(ob_res["rate_ist"].index), "Connects")

        if period in ("weekly", "monthly"):
            ws = wb.create_sheet("OB Heatmap"); ws.sheet_view.showGridLines = False
            sh = Sheet(ws); sh.title("Outbound — Connect-Rate Heatmap",
                                     "Weekday × time-of-day (EST). Greener = higher connect rate.")
            write_heatmap(sh, "Connect % by weekday and EST window", ob_raw, "EST Bucket",
                          ob_res.get("heat_rate"), ob_res.get("heat_tot"),
                          note="Target the greenest cells that also carry real dial volume.")

        ws = wb.create_sheet("OB Outcomes & Waste"); ws.sheet_view.showGridLines = False
        sh = Sheet(ws); sh.title("Outbound — Outcomes & Waste",
                                 "Where dials go when they don't connect — and when.")
        if wc is not None and len(wc) and tot:
            sh.section("Non-connect outcomes (dial leak, ranked)")
            data_start = sh.r + 1
            total_row = data_start + len(wc)
            rows = []
            for i, idx in enumerate(wc.index):
                er = data_start + i
                rows.append([idx, EQ(ob_raw.countif(ob_result_h, _q(idx))),
                             EQ(IFERR(f"B{er}/({total_dials})", "0"))])
            rows.append(["All non-connects", EQ(f"SUM(B{data_start}:B{total_row - 1})"),
                         EQ(IFERR(f"B{total_row}/({total_dials})", "0"))])
            sh.table(["Outcome", "Dials", "% of dials"], rows, pct_cols=(3,), num_cols=(2,),
                     widths=[26, 14, 14], total_row=True)
        write_outcome_matrix(sh, "Counts by EST window", ob_raw, "EST Bucket",
                             ob_result_h, ob_res["outcomes_est"])
        write_outcome_matrix_pct(sh, "Within-window % by EST window", ob_raw, "EST Bucket",
                                 ob_result_h, ob_res["outcomes_est"])

        ws = wb.create_sheet("OB Cost"); ws.sheet_view.showGridLines = False
        sh = Sheet(ws); sh.title("Outbound — Cost Analysis")
        has_cost = ob_res.get("total_cost") is not None and ob_cost_h
        deliv_h = ob_meta.get("delivery_cost"); link_h = ob_meta.get("linkback_cost")
        sh.section("Cost totals")
        if deliv_h and ob_raw.has(deliv_h):
            sh.kv("Delivery cost", EQ(ob_raw.sum(deliv_h)), '$#,##0.000')
        if link_h and ob_raw.has(link_h):
            sh.kv("Linkback cost", EQ(ob_raw.sum(link_h)), '$#,##0.000')
        tc_row = sh.r
        if has_cost:
            sh.kv("Total cost", EQ(ob_raw.sum(ob_cost_h)), '$#,##0.000', bold_val=True)
            sh.kv("Cost per dial", EQ(IFERR(f"B{tc_row}/({total_dials})", "0")), '$#,##0.0000')
            sh.kv("Cost per connect",
                  EQ(IFERR(f"B{tc_row}/{ob_raw.countif(H_ISCONNECT, _q('Yes'))}", "0")),
                  '$#,##0.0000', color=AMBER, bold_val=True)
            sh.blank()
            sh.section("Connected vs wasted spend")
            sh.kv("Spend that produced a connect",
                  EQ(ob_raw.sumif(H_ISCONNECT, _q("Yes"), ob_cost_h)),
                  '$#,##0.000', color=GREEN, bold_val=True)
            sh.kv("Spend on dials that didn't connect",
                  EQ(ob_raw.sumif(H_ISCONNECT, _q("No"), ob_cost_h)),
                  '$#,##0.000', color=RED, bold_val=True)
            sh.blank()
            cpc = ob_res.get("cost_per_connect_est")
            if cpc is not None and len(cpc):
                sh.section("Cost per connect by EST window (cheapest = best)")
                rows = []
                for b in BUCKET_ORDER:
                    if b in cpc.index:
                        spend = ob_raw.sumif("EST Bucket", _q(b), ob_cost_h)
                        conn = ob_raw.countifs([("EST Bucket", _q(b)), (H_ISCONNECT, _q("Yes"))])
                        rows.append([b, EQ(IFERR(f"({spend})/{conn}", '"-"'))])
                sh.table(["Time Window", "$ / Connect"], rows, money_cols=(2,), widths=[20, 16])
            cbr = ob_res.get("cost_by_result")
            if cbr is not None and len(cbr):
                sh.section("Cost by result")
                data_start = sh.r + 1
                total_row = data_start + len(cbr)
                rows = [[idx, EQ(ob_raw.sumif(ob_result_h, _q(idx), ob_cost_h))]
                        for idx in cbr.index]
                rows.append(["Total", EQ(f"SUM(B{data_start}:B{total_row - 1})")])
                sh.table(["Result", "Total Cost ($)"], rows, money_cols=(2,),
                         widths=[26, 18], total_row=True)

        write_ib_disposition_tabs(wb, ob_res.get("agentdisp"), ob_raw,
                                  talk_h=ob_meta.get("talk_dur"),
                                  prefix="OB", label="Outbound")

        _stream_period_tabs(wb, "OB", ob_res, period, "Answered Linkcalls", ob_raw)

    # ----- IB detail tabs ----- #
    if ib_res:
        ws = wb.create_sheet("IB Summary"); ws.sheet_view.showGridLines = False
        sh = Sheet(ws); sh.title("Inbound — Summary & Result Breakdown")
        rc = ib_res["result_counts"]; tot = ib_res["total"]
        sh.section("Result breakdown (all inbound)")
        _count_pct_table(sh, ["Result", "Calls", "% of inbound"],
                         [(idx, ib_raw.countif(ib_result_h, _q(idx))) for idx in rc.index],
                         (26, 14, 14))
        sh.section("Connectivity")
        total_ib = ib_raw.counta(ib_result_h)
        conn_row = sh.r
        sh.kv("Connected (Answered Linkcall + Answered)",
              EQ(ib_raw.countif(H_ISCONNECT, _q("Yes"))), "#,##0", color=GREEN, bold_val=True)
        sh.kv("Connectivity rate", EQ(IFERR(f"B{conn_row}/({total_ib})", "0")),
              "0.0%", bold_val=True)
        miss_row = sh.r
        sh.kv("Missed / not connected", EQ(ib_raw.countif(H_ISCONNECT, _q("No"))),
              "#,##0", color=RED, bold_val=True)
        sh.kv("Miss rate", EQ(IFERR(f"B{miss_row}/({total_ib})", "0")), "0.0%", color=RED)
        ib_talk_h = ib_meta.get("talk_dur")
        if ib_res.get("avg_talk") is not None and ib_talk_h and ib_raw.has(ib_talk_h):
            sh.kv("Avg agent talk time (sec)",
                  EQ(IFERR(f"AVERAGE({ib_raw.rng(ib_talk_h)})", "0")), "0.0")

        ws = wb.create_sheet("IB Best Time & Missed"); ws.sheet_view.showGridLines = False
        sh = Sheet(ws); sh.title("Inbound — Best Time & Missed",
                                 "Connect & miss rate by window — when inbound demand peaks and where it goes unanswered.")
        write_rate_window_table(sh, "Connect rate by EST window", ib_raw, "EST Bucket",
                                list(ib_res["rate_est"].index), "Connected",
                                note="Connect % here = share of inbound calls in that window that reached an agent.")
        write_distribution_table(sh, "Missed inbound by EST window", ib_raw, "EST Bucket",
                                 list(ib_res["miss_est"].index), "Missed Calls")
        write_distribution_table(sh, "Missed inbound by IST window (team time)", ib_raw,
                                 "IST Bucket", list(ib_res["miss_ist"].index), "Missed Calls")

        if period in ("weekly", "monthly"):
            ws = wb.create_sheet("IB Heatmap"); ws.sheet_view.showGridLines = False
            sh = Sheet(ws); sh.title("Inbound — Connect-Rate Heatmap",
                                     "Weekday × time-of-day (EST). Greener = higher connect rate.")
            write_heatmap(sh, "Connect % by weekday and EST window", ib_raw, "EST Bucket",
                          ib_res.get("heat_rate"), ib_res.get("heat_tot"),
                          note="Red cells with high volume = add agent coverage there.")

        ws = wb.create_sheet("IB Cost"); ws.sheet_view.showGridLines = False
        sh = Sheet(ws); sh.title("Inbound — Cost Analysis")
        ib_inb_h = ib_meta.get("inbound_cost"); ib_tot_h = ib_meta.get("total_cost")
        if ib_inb_h and ib_raw.has(ib_inb_h):
            sh.kv("Inbound cost", EQ(ib_raw.sum(ib_inb_h)), '$#,##0.000')
        if ib_tot_h and ib_raw.has(ib_tot_h):
            sh.kv("Total cost", EQ(ib_raw.sum(ib_tot_h)), '$#,##0.000', bold_val=True)
        if ib_res.get("total_talk") is not None and ib_talk_h and ib_raw.has(ib_talk_h):
            sh.kv("Total agent talk time (sec)", EQ(ib_raw.sum(ib_talk_h)), "#,##0")

        write_ib_disposition_tabs(wb, ib_res.get("agentdisp"), ib_raw, talk_h=ib_talk_h)

        _stream_period_tabs(wb, "IB", ib_res, period, "Connected Calls", ib_raw)

    # ----- embedded raw-data sheets (the source every formula points at) ----- #
    if ob_res:
        write_raw_sheet(wb, "OB_RawData", ob_raw_df)
    if ib_res:
        write_raw_sheet(wb, "IB_RawData", ib_raw_df)

    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Cross-client portfolio roll-up                                               #
# --------------------------------------------------------------------------- #
def client_metrics(client, period, date_label, ob_res, ib_res, drive_link=None,
                   period_start=None, period_end=None):
    """Flatten one client's results into a compact record for the portfolio
    roll-up + the trend history. Safe when a stream is missing.
    period_start / period_end (YYYY-MM-DD) anchor the record on the time axis
    so WoW/MoM/DoD comparisons can find the right prior period."""
    rec = {"client": client, "period": period, "date_label": date_label,
           "period_start": str(period_start) if period_start else None,
           "period_end": str(period_end) if period_end else None,
           "drive_link": drive_link, "ob": None, "ib": None}
    if ob_res:
        peaks = peak_windows(ob_res.get("rate_est"))
        rec["ob"] = {
            "dials": ob_res["total"],
            "connects": ob_res["answered_linkcalls"],
            "rate": ob_res["connectivity"],
            "dials_per_connect": ob_res.get("dials_per_connect"),
            "cost_per_connect": ob_res.get("cost_per_al"),
            "total_cost": ob_res.get("total_cost"),
            "best_window": peaks[0][0] if peaks else None,
            "best_window_rate": peaks[0][3] if peaks else None,
        }
        oad = ob_res.get("agentdisp")
        if oad:
            rec["ob"]["disp_total"] = oad["disp_total"]
            rec["ob"]["disp_connected"] = oad["disp_connected"]
            rec["ob"]["disp_connected_rate"] = oad["disp_connected_rate"]
            rec["ob"]["dispositions"] = {str(k): int(v)
                                         for k, v in oad["disp_counts"].items()}
            rec["ob"]["agents"] = {
                r["agent"]: {"handled": r["handled"], "connected": r["connected"],
                             "rate": r["rate"], "ptp": r["ptp"],
                             "payment": r["payment"], "avg_talk": r["avg_talk"]}
                for r in oad["agents"]}
    if ib_res:
        mb = best_bucket(ib_res.get("miss_est"))
        rec["ib"] = {
            "calls": ib_res["total"],
            "connected": ib_res["connected"],
            "rate": ib_res["connectivity"],
            "missed": ib_res["miss_total"],
            "miss_rate": ib_res.get("miss_rate", 0),
            "most_missed_window": mb[0] if mb[1] else None,
            "most_missed_count": mb[1],
        }
        # agent + disposition history (keyed maps -> per-key WoW/MoM trends)
        ad = ib_res.get("agentdisp")
        if ad:
            rec["ib"]["disp_total"] = ad["disp_total"]
            rec["ib"]["disp_connected"] = ad["disp_connected"]
            rec["ib"]["disp_connected_rate"] = ad["disp_connected_rate"]
            rec["ib"]["dispositions"] = {str(k): int(v)
                                         for k, v in ad["disp_counts"].items()}
            rec["ib"]["agents"] = {
                r["agent"]: {"handled": r["handled"], "connected": r["connected"],
                             "rate": r["rate"], "ptp": r["ptp"],
                             "payment": r["payment"], "avg_talk": r["avg_talk"]}
                for r in ad["agents"]}
    return rec


def _mv_strings(movement, client):
    """(Δ-connect-pts string, rank-change string) for one client in the roll-up."""
    m = (movement or {}).get(client, {})
    dp = m.get("delta_pts")
    if dp is None:
        ds = "—"
    else:
        arrow = "▲" if dp > 1e-12 else ("▬" if abs(dp) < 1e-12 else "▼")
        sign = "+" if dp > 0 else ("" if dp == 0 else "−")
        ds = f"{arrow} {sign}{abs(dp) * 100:.1f}"
    rc = m.get("rank_change")
    if rc is None:
        rs = "—"
    elif rc > 0:
        rs = f"↑{rc}"
    elif rc < 0:
        rs = f"↓{abs(rc)}"
    else:
        rs = "▬"
    return ds, rs


def build_portfolio(records, period, date_label, out_path, movement=None, alerts=None):
    """One-page cross-client roll-up. Ranks clients by outbound connect rate,
    totals the portfolio, and flags the strongest / weakest performers so
    leadership can see every account at a glance. When `movement` is provided
    (per-client prior connect rate + rank), adds WoW/MoM Δ and rank-change
    columns plus a 'Biggest mover' callout."""
    wb = Workbook()
    ws = wb.active; ws.title = "Portfolio"
    ws.sheet_view.showGridLines = False
    sh = Sheet(ws)
    sh.title(f"Portfolio Roll-Up — {len(records)} client(s)",
             f"{period.title()} · {date_label} · ranked by outbound connect rate · "
             f"generated {datetime.now():%Y-%m-%d %H:%M}")

    ob_recs = [r for r in records if r.get("ob")]
    ib_recs = [r for r in records if r.get("ib")]

    # ---- portfolio totals (KPI strip) ----
    tot_dials = sum(r["ob"]["dials"] for r in ob_recs)
    tot_conn = sum(r["ob"]["connects"] for r in ob_recs)
    tot_spend = sum(r["ob"]["total_cost"] or 0 for r in ob_recs)
    port_rate = pct(tot_conn, tot_dials)
    port_cpc = pct(tot_spend, tot_conn) if tot_spend else 0
    if ob_recs:
        sh.section("Outbound — portfolio totals")
        items = [("Clients", len(ob_recs), "#,##0", "FFFFFF"),
                 ("Total dials", tot_dials, "#,##0", "FFFFFF"),
                 ("Total connects", tot_conn, "#,##0", "C6E0B4"),
                 ("Blended connect rate", port_rate, "0.0%", "C6E0B4")]
        if tot_spend:
            items.append(("Total spend", tot_spend, '$#,##0', "FFFFFF"))
            items.append(("Blended $/connect", port_cpc, '$#,##0.000', "FFE699"))
        sh.kpi_row(items)
        sh.blank()

    # ---- per-client outbound ranking ----
    has_move = bool(movement)
    if ob_recs:
        ranked = sorted(ob_recs, key=lambda r: r["ob"]["rate"], reverse=True)
        basis = {"weekly": "WoW", "monthly": "MoM"}.get(period, "DoD")
        sh.section("Outbound by client (best connect rate first)")
        rows = []
        for r in ranked:
            o = r["ob"]
            dpc = round(o["dials_per_connect"], 1) if o.get("dials_per_connect") else "-"
            cpc = o["cost_per_connect"] if o.get("cost_per_connect") is not None else "-"
            row = [r["client"], o["dials"], o["connects"], o["rate"]]
            if has_move:
                ds, rs = _mv_strings(movement, r["client"])
                row += [ds, rs]
            row += [dpc, cpc, o.get("best_window") or "-"]
            rows.append(row)
        trow = ["PORTFOLIO", tot_dials, tot_conn, port_rate]
        if has_move:
            trow += ["", ""]
        trow += [(round(tot_dials / tot_conn, 1) if tot_conn else "-"),
                 (port_cpc if tot_spend else "-"), ""]
        rows.append(trow)
        if has_move:
            headers = ["Client", "Dials", "Connects", "Connect %",
                       f"Δ Connect ({basis}, pts)", "Rank Δ",
                       "Dials / Connect", "$ / Connect", "Best Window (EST)"]
            sh.table(headers, rows, pct_cols=(4,), num_cols=(2, 3), money_cols=(8,),
                     widths=[22, 11, 11, 11, 16, 9, 14, 12, 20], total_row=True)
        else:
            headers = ["Client", "Dials", "Connects", "Connect %",
                       "Dials / Connect", "$ / Connect", "Best Window (EST)"]
            sh.table(headers, rows, pct_cols=(4,), num_cols=(2, 3), money_cols=(6,),
                     widths=[22, 12, 12, 12, 14, 12, 20], total_row=True)

        best, worst = ranked[0], ranked[-1]
        sh.kv("Top performer", f"{best['client']} — {best['ob']['rate']:.1%} connect",
              color=GREEN, bold_val=True)
        if worst["client"] != best["client"]:
            sh.kv("Needs attention", f"{worst['client']} — {worst['ob']['rate']:.1%} connect",
                  color=RED, bold_val=True)
        if has_move:
            movers = [(c, m["delta_pts"]) for c, m in movement.items()
                      if m.get("delta_pts") is not None]
            if movers:
                bc, bd = max(movers, key=lambda x: abs(x[1]))
                arrow = "▲" if bd >= 0 else "▼"
                sh.kv(f"Biggest mover ({basis})",
                      f"{bc} — {arrow} {abs(bd) * 100:.1f} pts connect rate",
                      color=(GREEN if bd >= 0 else RED), bold_val=True)
        sh.blank()

    # ---- per-client inbound ----
    if ib_recs:
        ranked_ib = sorted(ib_recs, key=lambda r: r["ib"]["miss_rate"])
        sh.section("Inbound by client (lowest miss rate first)")
        t_calls = sum(r["ib"]["calls"] for r in ib_recs)
        t_conn = sum(r["ib"]["connected"] for r in ib_recs)
        t_miss = sum(r["ib"]["missed"] for r in ib_recs)
        rows = []
        for r in ranked_ib:
            b = r["ib"]
            rows.append([r["client"], b["calls"], b["connected"], b["rate"],
                         b["missed"], b["miss_rate"], b.get("most_missed_window") or "-"])
        rows.append(["PORTFOLIO", t_calls, t_conn, pct(t_conn, t_calls),
                     t_miss, pct(t_miss, t_calls), ""])
        sh.table(["Client", "Calls", "Connected", "Connect %", "Missed",
                  "Miss %", "Most-Missed Window"], rows,
                 pct_cols=(4, 6), num_cols=(2, 3, 5),
                 widths=[22, 12, 12, 12, 10, 10, 20], total_row=True)
        sh.blank()

    # ---- portfolio alerts ----
    if alerts:
        sh.section("Portfolio alerts")
        for a in alerts:
            marker = {"CRITICAL": "🔴", "WARN": "🟠", "INFO": "🔵"}.get(a["severity"], "•")
            sh.kv(f"{marker} {a['metric']}", a["message"],
                  color=SEV_TEXT.get(a["severity"], GREY),
                  bold_val=(a["severity"] == "CRITICAL"))
        sh.blank()

    sh.section("Notes")
    sh.note("• Ranked by connect performance so leadership can triage accounts fastest-to-slowest.")
    sh.note("• Each client's full report (with best-time, heatmap, cost drill-downs) lives in its own folder.")
    sh.note("• 'Best Window' = highest connect-rate 2-hour EST window carrying real dial volume.")
    ws.column_dimensions["A"].width = 24
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 15
    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Filename parsing                                                             #
# --------------------------------------------------------------------------- #
def parse_filename(path):
    stem = Path(path).stem
    parts = re.split(r"[_\s]+", stem)
    client, period, datelabel = None, None, None
    for p in parts:
        if p.lower() in ("daily", "weekly", "monthly"):
            period = p.lower()
    dates = re.findall(r"\d{4}-\d{2}(?:-\d{2})?", stem)
    if dates:
        datelabel = " to ".join(dates) if len(dates) > 1 else dates[0]
    if parts and parts[0].lower() not in ("daily", "weekly", "monthly"):
        client = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", parts[0])
    return client, period, datelabel


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--out", default=None)
    ap.add_argument("--client", default=None)
    ap.add_argument("--period", default=None, choices=["daily", "weekly", "monthly"])
    args = ap.parse_args()

    fc, fp, fd = parse_filename(args.input)
    client = args.client or fc or "Client"
    period = args.period or fp or "daily"

    ob, ib = load(args.input)
    if ob is None and ib is None:
        print("ERROR: could not find OB or IB sheets.", file=sys.stderr)
        sys.exit(1)

    # date label from data if not in filename
    date_label = fd
    if not date_label:
        dts = []
        for df in (ob, ib):
            if df is not None:
                dts += [d for d in df["EST Date"].dropna().tolist()]
        if dts:
            lo, hi = min(dts), max(dts)
            date_label = str(lo) if lo == hi else f"{lo} to {hi}"

    ob_res = analyze_ob(ob) if ob is not None else None
    ib_res = analyze_ib(ib) if ib is not None else None

    out = args.out or str(Path(args.input).with_name(
        f"{client.replace(' ', '')}_{period.title()}_{(fd or date_label or 'report').replace(' ', '')}_REPORT.xlsx"))
    build_workbook(client, period, date_label or "—", ob_res, ib_res, out)
    print(f"OK -> {out}")
    if ob_res:
        print(f"  OB: {ob_res['total']} dials, {ob_res['answered_linkcalls']} answered linkcalls "
              f"({ob_res['connectivity']:.1%})")
    if ib_res:
        print(f"  IB: {ib_res['total']} inbound, {ib_res['connected']} connected "
              f"({ib_res['connectivity']:.1%}), {ib_res['miss_total']} missed")


if __name__ == "__main__":
    main()
